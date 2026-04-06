# coolWIRE_main.py
# °coolWIRE v2.0 – Kälteplanungstool
# (c) coolsulting e.U. | Michael Schäpers | Linz, Austria
# Port: 8580

import streamlit as st
import pandas as pd
import json
import math
from datetime import datetime

from modules.auth import (
    ist_eingeloggt, ist_admin, get_display_name, logout,
    zeige_login, zeige_nutzerverwaltung
)
from modules.kuehlstellen import (
    KOMPONENTEN, KOMPONENTEN_GRUPPEN, TEMP_BEREICHE, STEUERUNGSSYSTEME,
    neue_kuehlstelle, kopiere_kuehlstelle,
    berechne_kuehlstellen_kabel, gruppiere_nach_kreis, kreis_zusammenfassung,
    get_temp_bereich_info
)
from modules.ki_import import (
    api_verfuegbar, analysiere_pdf, analysiere_dwg,
    analysiere_freitext, ki_ergebnis_zu_kuehlstellen, get_erkannte_kreise,
    merge_kuehlstellen,
    analysiere_datenblatt_pdf, analysiere_datenblatt_bild, analysiere_datenblatt_excel
)
from modules.datenbank import (
    lade_datenbank, speichere_datenbank, get_regler_liste,
    get_regler_optionen, add_regler, delete_regler,
    get_theme, speichere_theme, get_theme_presets, theme_zu_css, DEFAULT_DB,
    schreibe_config_toml, get_geraete, get_geraet_optionen, add_geraet,
    delete_geraet, geraet_zu_kuehlstelle, GERAET_TYPEN, DEFAULT_GERAET
)
from modules.kabelliste import erzeuge_kabelliste, kabelliste_zusammenfassung
from modules.calculation_kabel import (
    KABEL_MATRIX, berechne_leitungsquerschnitt, exportiere_gesamtliste
)

st.set_page_config(page_title="°coolWIRE v2.0", page_icon="⚡", layout="wide",
                   initial_sidebar_state="expanded")

# DB und Theme direkt beim Start laden (vor Login und CSS)
_db_early = lade_datenbank()
_theme = get_theme(_db_early)

# Dynamic CSS with theme
_cb   = _theme.get('primary_color','#36A9E1')
_cd   = _theme.get('dark_color','#3C3C3B')
_cl   = _theme.get('light_bg','#F4F8FC')
_cbrd = _theme.get('border_color','#D0E8F5')
_ca   = _theme.get('accent_color','#0078B8')
_fnt  = _theme.get('font_family','DM Sans')
_fsz  = _theme.get('font_size_base','14px')
_rad  = _theme.get('border_radius','10px')
_gs   = _theme.get('header_gradient_start','#1a6fa8')
_ge   = _theme.get('header_gradient_end','#5bc4f5')

st.markdown(f"""
<style>
@import url('https://fonts.googleapis.com/css2?family={_fnt.replace(" ","+")}:wght@300;400;500;600;700;800&display=swap');
@import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@300;400;500;600;700;800&display=swap');
:root{{--cb:{_cb};--cd:{_cd};--cl:{_cl};--cborder:{_cbrd};--ca:{_ca};--ok:#27AE60;--warn:#E67E22;--err:#E74C3C;--radius:{_rad};--grad-start:{_gs};--grad-end:{_ge};}}
html,body,[class*="css"]{{font-family:'{_fnt}',sans-serif;font-size:{_fsz};color:var(--cd);}}
.hdr{{background:linear-gradient(135deg,{_gs} 0%,{_cb} 60%,{_ge} 100%);padding:1.2rem 1.8rem;border-radius:var(--radius);margin-bottom:1.2rem;display:flex;align-items:center;justify-content:space-between;box-shadow:0 4px 20px rgba(54,169,225,0.25);}}
.hdr h1{{color:white;font-size:1.7rem;font-weight:800;margin:0;letter-spacing:-0.5px;}}
.hdr p{{color:rgba(255,255,255,0.82);font-size:0.8rem;margin:0;}}
.hdr-right{{color:rgba(255,255,255,0.9);font-size:0.83rem;text-align:right;}}
.vtag{{background:rgba(255,255,255,0.22);color:white;padding:2px 10px;border-radius:20px;font-size:0.72rem;font-weight:700;}}
.stTabs [data-baseweb="tab-list"]{{gap:6px;background:var(--cl);padding:7px;border-radius:10px;border:1px solid var(--cborder);}}
.stTabs [data-baseweb="tab"]{{border-radius:8px;padding:7px 18px;font-weight:600;font-size:0.88rem;background:transparent;border:none;}}
.stTabs [aria-selected="true"]{{background:var(--cb) !important;color:white !important;}}
.sec{{font-size:1rem;font-weight:700;color:var(--cd);border-left:4px solid var(--cb);padding-left:11px;margin:1.1rem 0 0.7rem 0;}}
.hint-box{{background:#FEF9E7;border-left:4px solid #F39C12;padding:0.6rem 0.9rem;border-radius:0 8px 8px 0;font-size:0.83rem;margin:0.4rem 0;}}
.info-box{{background:#EBF5FB;border-left:4px solid var(--cb);padding:0.6rem 0.9rem;border-radius:0 8px 8px 0;font-size:0.83rem;margin:0.4rem 0;}}
.ks-card{{background:white;border:1.5px solid var(--cborder);border-radius:10px;padding:0.9rem 1.1rem;margin-bottom:0.5rem;box-shadow:0 2px 8px rgba(0,0,0,0.04);}}
.komp-box{{background:var(--cl);border:1px solid var(--cborder);border-radius:8px;padding:0.7rem 0.9rem;margin-bottom:0.4rem;}}
.sb-logo{{background:var(--cb);color:white;padding:0.7rem 1rem;border-radius:10px;text-align:center;font-weight:800;font-size:1.05rem;letter-spacing:-0.5px;margin-bottom:0.8rem;}}
.sb-user{{background:var(--cl);border:1px solid var(--cborder);border-radius:8px;padding:0.55rem 0.75rem;margin-bottom:0.5rem;font-size:0.8rem;}}
.stDownloadButton>button{{background:var(--cb) !important;color:white !important;border:none !important;border-radius:8px !important;font-weight:600 !important;}}
.stButton>button{{border-radius:8px !important;font-weight:600 !important;}}
.footer{{margin-top:2rem;padding:0.6rem 1rem;background:var(--cl);border-top:2px solid var(--cborder);font-size:0.72rem;color:#aaa;text-align:center;border-radius:0 0 10px 10px;}}
</style>
""", unsafe_allow_html=True)

# LOGIN
if not ist_eingeloggt():
    zeige_login()
    st.stop()

# SESSION STATE
def _init():
    defs = {
        "_projekt_key": "init",
        "projekt": {"name":"","nummer":"","kunde":"","standort":"","land":"AT",
                    "bearbeiter":get_display_name(),"erstellt":datetime.now().strftime("%d.%m.%Y"),"notizen":""},
        "kuehlstellen": [],
        "aktive_ks_idx": 0,
        "steuerung": None,
        "ki_ergebnis": None,
        "qs_result": None,
        "db": None,
        "globaler_sk": False,
        "globaler_sk_bezeichnung": "SK-01 Maschinenraum",
        "maschinenstandorte": [
            {"id": "ms1", "anlage_typ": "Verbundanlage", "standort_maschine": "Maschinenraum", "etage_maschine": "EG", "standort_verfluessiger": "", "etage_verfluessiger": "", "waereruckgewinnung": False, "beschreibung": "", "kreise": []}
        ],
        "ki_rueckfrage": None,
        "haccp_module": [],
        "kabelliste_cache": None,
        "global_tueren": 1,
        "global_licht": True,
        "global_bewegungsmelder": True,
        "global_tuer_kontakt": True,
        "global_haccp": True,
    }
    for k,v in defs.items():
        if k not in st.session_state:
            st.session_state[k] = v
_init()

# DB LADEN
if st.session_state.db is None:
    st.session_state.db = lade_datenbank()
_db = st.session_state.db

# THEME AUS DB LADEN
_theme = get_theme(_db)

# HEADER
rolle = "🔑 Admin" if ist_admin() else "🤝 Partner"
st.markdown(f"""<div class="hdr">
  <div><div style="display:flex;align-items:center;gap:10px;"><h1>⚡ °coolWIRE</h1><span class="vtag">v2.0</span></div><p>Kabelplanungstool · coolsulting e.U. · Linz</p></div>
  <div class="hdr-right">👤 <strong>{get_display_name()}</strong><br><span style="opacity:0.75;">{rolle}</span></div>
</div>""", unsafe_allow_html=True)

# SIDEBAR
with st.sidebar:
    st.markdown('<div class="sb-logo">°coolsulting</div>', unsafe_allow_html=True)
    st.markdown(f'<div class="sb-user">👤 <strong>{get_display_name()}</strong><br><span style="color:#888;">{rolle}</span></div>', unsafe_allow_html=True)
    if st.button("🚪 Abmelden", use_container_width=True):
        logout(); st.rerun()
    st.markdown("---")
    p = st.session_state.projekt
    if p.get("name"):
        st.markdown(f"**📁 {p['name']}**")
        st.caption(f"Nr. {p.get('nummer','–')} · {p.get('kunde','–')}")
        st.metric("Kühlstellen", len(st.session_state.kuehlstellen))
    st.markdown("---")
    if st.session_state.kuehlstellen:
        exp = {"version":"2.0","projekt":st.session_state.projekt,
               "kuehlstellen":st.session_state.kuehlstellen,
               "steuerung": st.session_state.steuerung,
               "maschinenstandorte": st.session_state.get("maschinenstandorte",[]),
               "globaler_sk": st.session_state.get("globaler_sk", False),
               "globaler_sk_bezeichnung": st.session_state.get("globaler_sk_bezeichnung",""),
               "global_tuer_kontakt": st.session_state.get("global_tuer_kontakt", True),
               "global_tueren": st.session_state.get("global_tueren", 1),
               "global_licht": st.session_state.get("global_licht", True),
               "global_bewegungsmelder": st.session_state.get("global_bewegungsmelder", True),
               "global_haccp": st.session_state.get("global_haccp", True),
               "haccp_module": st.session_state.get("haccp_module", []),
               "gespeichert":datetime.now().isoformat()}
        st.download_button("💾 Projekt speichern",
            data=json.dumps(exp,ensure_ascii=False,indent=2).encode("utf-8"),
            file_name=f"coolWIRE_{p.get('nummer','projekt')}.json",
            mime="application/json", use_container_width=True)
    up = st.file_uploader("📂 Projekt laden", type=["json"], label_visibility="collapsed")
    if up:
        try:
            data = json.load(up)
            p_data = data.get("projekt", {})
            st.session_state.projekt = p_data
            # Widget-Keys LÖSCHEN damit Streamlit sie neu rendert
            for field in ["name","nummer","kunde","standort","bearbeiter","erstellt","notizen"]:
                key = "proj_" + field
                if key in st.session_state:
                    del st.session_state[key]
                st.session_state[key] = p_data.get(field, "")
            st.session_state.kuehlstellen      = data.get("kuehlstellen", [])
            st.session_state._projekt_key      = str(int(datetime.now().timestamp() * 1000))
            st.session_state.steuerung         = data.get("steuerung")
            st.session_state.maschinenstandorte= data.get("maschinenstandorte", st.session_state.maschinenstandorte)
            st.session_state.globaler_sk       = data.get("globaler_sk", False)
            st.session_state.globaler_sk_bezeichnung = data.get("globaler_sk_bezeichnung", "SK-01")
            st.session_state.global_tuer_kontakt = data.get("global_tuer_kontakt", True)
            st.session_state.global_tueren     = data.get("global_tueren", 1)
            st.session_state.global_licht      = data.get("global_licht", True)
            st.session_state.global_bewegungsmelder = data.get("global_bewegungsmelder", True)
            st.session_state.global_haccp      = data.get("global_haccp", True)
            st.session_state.haccp_module      = data.get("haccp_module", [])
            st.session_state.kabelliste_cache  = None
            st.success(f"✅ Projekt geladen: {p_data.get('name','?')} · {len(st.session_state.kuehlstellen)} Kühlstellen")
            # KEIN st.rerun() – Streamlit rendert die Widgets im selben Durchlauf neu
            # st.rerun() würde den Widget-Cache nicht leeren
        except Exception as e:
            st.error(str(e))
    st.markdown("---")
    st.caption(f"🕐 {datetime.now().strftime('%d.%m.%Y %H:%M')}")

# TABS
tl = ["📁 1 · Projekt","❄️ 2 · Kühlstellen","🧮 3 · Steuerung","🧵 4 · Kabelplanung","╔ 5 · Rohrnetz","📄 6 · Doku"]
if ist_admin(): tl.append("🔑 7 · Admin")
tabs = st.tabs(tl)
t1,t2,t3,t4,t5_rohr,t5,t6_admin = tabs[0],tabs[1],tabs[2],tabs[3],tabs[4],tabs[5],tabs[6] if ist_admin() else None
t6 = t6_admin

# ============================================================
# TAB 1 – PROJEKT
# ============================================================
with t1:
    st.markdown('<div class="sec">📁 Projektdaten</div>', unsafe_allow_html=True)
    p = st.session_state.projekt

    # Projektdaten anzeigen – immer direkt aus session_state.projekt
    c1,c2 = st.columns(2, gap="large")
    with c1:
        neu_name      = st.text_input("Projektname *",   value=p.get("name",""),      placeholder="z.B. Supermarkt Müller Linz")
        neu_nummer    = st.text_input("Projektnummer *", value=p.get("nummer",""),    placeholder="2025-042")
        neu_kunde     = st.text_input("Kunde",           value=p.get("kunde",""))
        neu_standort  = st.text_input("Standort",        value=p.get("standort",""))
    with c2:
        _land_opts = ["AT","DE","HR"]
        _land_cur  = p.get("land","AT")
        _land_idx  = _land_opts.index(_land_cur) if _land_cur in _land_opts else 0
        neu_land      = st.selectbox("Normland", _land_opts,
                            format_func=lambda x:{"AT":"🇦🇹 Österreich","DE":"🇩🇪 Deutschland","HR":"🇭🇷 Kroatien"}[x],
                            index=_land_idx)
        neu_bearb     = st.text_input("Bearbeiter",      value=p.get("bearbeiter",""))
        neu_datum     = st.text_input("Datum",           value=p.get("erstellt",""))
        neu_notizen   = st.text_area("Notizen",          value=p.get("notizen",""),   height=100)

    # Werte sofort in session_state schreiben (ohne Button)
    st.session_state.projekt.update({
        "name": neu_name, "nummer": neu_nummer, "kunde": neu_kunde,
        "standort": neu_standort, "land": neu_land, "bearbeiter": neu_bearb,
        "erstellt": neu_datum, "notizen": neu_notizen
    })

    st.markdown("---")
    st.markdown('<div class="sec">🏭 Kälteanlagen & Standorte</div>', unsafe_allow_html=True)
    st.caption("Definiere alle Kälteanlagen, Aggregate und deren Standorte. Maschine und Verflüssiger können an unterschiedlichen Orten stehen.")

    ANLAGE_TYPEN = [
        "Verbundanlage",
        "Außenunit / Verflüssigungseinheit (Split)",
        "Rumpfaggregat + getrennter Verflüssiger",
        "Monoblock / Kompaktaggregat",
        "Gaskühlereinheit (CO₂)",
        "Wärmepumpe",
        "Prozesskältemaschine / Chiller",
        "Sonstiges",
    ]
    STANDORT_OPTIONEN = [
        "Dach", "Außenwand", "Balkon / Terrasse",
        "Maschinenraum EG", "Maschinenraum OG", "Maschinenraum KG / Keller",
        "Technikraum", "Außenaufstellung (ebenerdig)",
        "Container", "Sonstiges",
    ]

    _ms_default = [{"id":"ms1","anlage_typ":"Verbundanlage",
                    "standort_maschine":"Maschinenraum EG","etage_maschine":"EG",
                    "standort_verfluessiger":"Dach","etage_verfluessiger":"Dach",
                    "verfluessiger_getrennt":False,"waermerueckgewinnung":False,
                    "beschreibung":"","kreise":[]}]
    ms_list = st.session_state.get("maschinenstandorte", _ms_default)

    for ms_i, ms in enumerate(ms_list):
        with st.expander(
            f"Anlage {ms_i+1} – {ms.get('anlage_typ','?')} · "
            f"{ms.get('standort_maschine','?')} "
            f"{'/ Verflüssiger: ' + ms.get('standort_verfluessiger','') if ms.get('verfluessiger_getrennt') else ''}",
            expanded=True):

            ms_r1c1, ms_r1c2, ms_r1c3 = st.columns([2,2,1])
            with ms_r1c1:
                anlage_idx = ANLAGE_TYPEN.index(ms.get("anlage_typ", ANLAGE_TYPEN[0])) if ms.get("anlage_typ") in ANLAGE_TYPEN else 0
                ms["anlage_typ"] = st.selectbox("Anlagentyp",
                    ANLAGE_TYPEN, index=anlage_idx, key=f"ms_typ_{ms_i}",
                    help="Art der Kälteanlage")
            with ms_r1c2:
                ms["beschreibung"] = st.text_input("Bezeichnung / Beschreibung",
                    value=ms.get("beschreibung",""), key=f"ms_desc_{ms_i}",
                    placeholder="z.B. Verbund NK+TK 3 Verdichter, Außenunit Split KR-01")
            with ms_r1c3:
                ms["waermerueckgewinnung"] = st.checkbox("♻️ Wärmerückgewinnung",
                    value=ms.get("waermerueckgewinnung", False), key=f"ms_wrg_{ms_i}",
                    help="Wärmerückgewinnung vorhanden – zusätzliche Leitungen zum WRG-Wärmetauscher")

            ms_r2c1, ms_r2c2, ms_r2c3, ms_r2c4 = st.columns([2,1,2,1])
            with ms_r2c1:
                sm_idx = STANDORT_OPTIONEN.index(ms.get("standort_maschine", STANDORT_OPTIONEN[3])) if ms.get("standort_maschine") in STANDORT_OPTIONEN else 3
                ms["standort_maschine"] = st.selectbox("📍 Standort Maschine / Verdichter",
                    STANDORT_OPTIONEN, index=sm_idx, key=f"ms_sma_{ms_i}",
                    help="Wo steht die Kältemaschine / der Verdichter?")
            with ms_r2c2:
                ms["etage_maschine"] = st.text_input("Etage",
                    value=ms.get("etage_maschine",""), key=f"ms_ema_{ms_i}",
                    placeholder="EG, 1.OG...")

            ms["verfluessiger_getrennt"] = st.checkbox(
                "Verflüssiger / Gaskühler an anderem Standort",
                value=ms.get("verfluessiger_getrennt", False), key=f"ms_vgt_{ms_i}",
                help="z.B. Rumpfaggregat im Keller, Verflüssiger auf dem Dach")

            if ms.get("verfluessiger_getrennt"):
                with ms_r2c3:
                    sv_idx = STANDORT_OPTIONEN.index(ms.get("standort_verfluessiger", STANDORT_OPTIONEN[0])) if ms.get("standort_verfluessiger") in STANDORT_OPTIONEN else 0
                    ms["standort_verfluessiger"] = st.selectbox("📍 Standort Verflüssiger / Gaskühler",
                        STANDORT_OPTIONEN, index=sv_idx, key=f"ms_svf_{ms_i}",
                        help="Wo steht der Verflüssiger oder Gaskühler?")
                with ms_r2c4:
                    ms["etage_verfluessiger"] = st.text_input("Etage",
                        value=ms.get("etage_verfluessiger",""), key=f"ms_evf_{ms_i}",
                        placeholder="Dach...")

            if len(ms_list) > 1:
                if st.button(f"🗑️ Anlage {ms_i+1} entfernen", key=f"ms_del_{ms_i}"):
                    ms_list.pop(ms_i)
                    st.session_state.maschinenstandorte = ms_list
                    st.rerun()

        ms_list[ms_i] = ms

    if st.button("➕ Weitere Kälteanlage / Aggregat hinzufügen"):
        ms_list.append({"id": f"ms{len(ms_list)+1}",
                        "anlage_typ": "Außenunit / Verflüssigungseinheit (Split)",
                        "standort_maschine": "Dach", "etage_maschine": "Dach",
                        "standort_verfluessiger": "", "etage_verfluessiger": "",
                        "verfluessiger_getrennt": False, "waermerueckgewinnung": False,
                        "beschreibung": "", "kreise": []})
        st.rerun()
    st.session_state.maschinenstandorte = ms_list

    st.markdown("---")
    st.markdown('<div class="sec">🔌 Schaltschrank-Konfiguration (global)</div>', unsafe_allow_html=True)
    st.caption("Globale Einstellung für alle Kühlstellen – kann pro Kühlstelle überschrieben werden.")
    gsk_col1, gsk_col2 = st.columns(2)
    with gsk_col1:
        st.session_state.globaler_sk = st.radio(
            "Schaltschrank-Typ",
            [False, True],
            format_func=lambda x: "🔌 Einzelregler an jeder Kühlstelle" if not x else "🏭 Zentraler Schaltschrank",
            index=1 if st.session_state.get("globaler_sk") else 0,
            horizontal=False,
            help="Zentraler Schaltschrank: alle Regler in einem Schrank, Bedienteil / Display optional an der Zelle"
        )
    with gsk_col2:
        if st.session_state.globaler_sk:
            st.session_state.globaler_sk_bezeichnung = st.text_input(
                "Bezeichnung Zentralschaltschrank",
                value=st.session_state.get("globaler_sk_bezeichnung","SK-01 Maschinenraum"),
                placeholder="z.B. SK-01 Maschinenraum EG",
                help="Erscheint in der Kabelliste und Dokumentation"
            )
            st.markdown('<div style="background:#EBF5FB;border-left:3px solid #36A9E1;padding:0.5rem 0.8rem;border-radius:0 8px 8px 0;font-size:0.83rem;">'
                        '💡 Regler im zentralen Schaltschrank, Bedienteil MTM / Gassensoranzeige / Notruf-Taster trotzdem an der Kühlstelle möglich – pro Kühlstelle einstellbar.</div>',
                        unsafe_allow_html=True)

    st.markdown("---")
    st.markdown('<div class="sec">🔧 Standard-Komponenten (global für alle Kühlstellen)</div>', unsafe_allow_html=True)
    st.caption("Diese Einstellungen gelten als Standard für alle neuen Kühlstellen – pro Kühlstelle in Tab 2 überschreibbar.")

    gc1, gc2, gc3, gc4, gc5 = st.columns(5)
    with gc1:
        st.session_state.global_tuer_kontakt = st.checkbox(
            "🚪 Türkontakt",
            value=st.session_state.get("global_tuer_kontakt", True),
            help="Türkontakt standardmäßig aktiv – Anzahl Türen wird pro Kühlstelle/Raum eingestellt")
        st.caption("Anzahl Türen → pro Kühlstelle")
    with gc2:
        st.session_state.global_licht = st.checkbox(
            "💡 Innenbeleuchtung",
            value=st.session_state.get("global_licht", True),
            help="Innenbeleuchtung standardmäßig aktiv")
    with gc3:
        st.session_state.global_bewegungsmelder = st.checkbox(
            "👁️ Bewegungsmelder",
            value=st.session_state.get("global_bewegungsmelder", True),
            help="Bewegungsmelder standardmäßig aktiv")
    with gc4:
        st.session_state.global_haccp = st.checkbox(
            "🌡️ HACCP-Aufzeichnung",
            value=st.session_state.get("global_haccp", True),
            help="HACCP Temperaturaufzeichnung standardmäßig aktiv – Fühlertyp pro Kühlstelle editierbar")
    with gc5:
        if st.button("🔄 Globale Standards auf alle Kühlstellen anwenden",
                     use_container_width=True,
                     help="Überschreibt die Einstellungen aller bestehenden Kühlstellen mit den globalen Werten"):
            from modules.kuehlstellen import KOMPONENTEN
            for ks_item in st.session_state.kuehlstellen:
                if "komponenten" not in ks_item: ks_item["komponenten"] = {}
                # Türkontakt
                if "tuer_kontakt" in KOMPONENTEN:
                    if "tuer_kontakt" not in ks_item["komponenten"]:
                        ks_item["komponenten"]["tuer_kontakt"] = {"aktiv": False, "parameter": {}, "ki_erkannt": False}
                    ks_item["komponenten"]["tuer_kontakt"]["aktiv"] = st.session_state.global_tuer_kontakt
                    # Türanzahl bleibt pro Kühlstelle – hier nicht überschreiben
                # Licht
                if "innenbeleuchtung" in KOMPONENTEN:
                    if "innenbeleuchtung" not in ks_item["komponenten"]:
                        ks_item["komponenten"]["innenbeleuchtung"] = {"aktiv": False, "parameter": {}, "ki_erkannt": False}
                    ks_item["komponenten"]["innenbeleuchtung"]["aktiv"] = st.session_state.global_licht
                # Bewegungsmelder
                if "bewegungsmelder" in KOMPONENTEN:
                    if "bewegungsmelder" not in ks_item["komponenten"]:
                        ks_item["komponenten"]["bewegungsmelder"] = {"aktiv": False, "parameter": {}, "ki_erkannt": False}
                    ks_item["komponenten"]["bewegungsmelder"]["aktiv"] = st.session_state.global_bewegungsmelder
                # HACCP
                if "haccp_aufzeichnung" in KOMPONENTEN:
                    if "haccp_aufzeichnung" not in ks_item["komponenten"]:
                        ks_item["komponenten"]["haccp_aufzeichnung"] = {"aktiv": False, "parameter": {}, "ki_erkannt": False}
                    ks_item["komponenten"]["haccp_aufzeichnung"]["aktiv"] = st.session_state.get("global_haccp", True)
            st.success(f"✅ Standards auf {len(st.session_state.kuehlstellen)} Kühlstelle(n) angewendet!")
            st.rerun()

    st.markdown("---")
    st.markdown('<div class="sec">📊 Projektstatus</div>', unsafe_allow_html=True)
    ks_list = st.session_state.kuehlstellen
    c1,c2,c3,c4 = st.columns(4)
    c1.metric("Kühlstellen", len(ks_list))
    c2.metric("Direkt", sum(1 for k in ks_list if k.get("lieferumfang")=="direkt"))
    c3.metric("Extern", sum(1 for k in ks_list if k.get("lieferumfang")=="extern"))
    c4.metric("Steuerung", st.session_state.steuerung or "–")

    if ks_list:
        st.markdown('<div class="sec">🌡️ Kreisübersicht & Zusammenfassung</div>', unsafe_allow_html=True)
        kreise_info = kreis_zusammenfassung(ks_list)

        for ki in kreise_info:
            farbe = ki["farbe"]
            st.markdown(f"""
            <div style="border-left:5px solid {farbe};background:white;border-radius:0 10px 10px 0;
                        padding:0.8rem 1rem;margin-bottom:0.6rem;box-shadow:0 2px 6px rgba(0,0,0,0.05);">
                <div style="display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:0.5rem;">
                    <div>
                        <span style="font-size:1rem;font-weight:700;color:{farbe};">
                            Kreis {ki["kreis_nr"]} – {ki["temp_label"]}
                        </span>
                        &nbsp;·&nbsp;
                        <span style="color:#555;font-size:0.88rem;">
                            Verdampfung <strong>{ki["verdampfung_c"]}°C</strong>
                        </span>
                    </div>
                    <div style="display:flex;gap:1rem;font-size:0.85rem;">
                        <span>❄️ <strong>{ki["anzahl_gesamt"]}</strong> Kühlstellen</span>
                        <span>⚡ <strong>{ki["leistung_kw_gesamt"]} kW</strong> gesamt</span>
                        <span>🔧 <strong>{ki["anzahl_direkt"]}</strong> Direkt</span>
                        <span>🔄 <strong>{ki["anzahl_extern"]}</strong> Extern</span>
                    </div>
                </div>
                <div style="margin-top:0.4rem;font-size:0.82rem;color:#666;">
                    {"🏭 " + " | ".join(ki["maschinenstandorte"]) if ki["maschinenstandorte"] and ki["maschinenstandorte"] != ["–"] else ""}
                    {"&nbsp;·&nbsp;" if ki["maschinenstandorte"] and ki["maschinenstandorte"] != ["–"] and ki["kuehlstellen_alle"] else ""}
                    {" · ".join(k["name"] for k in ki["kuehlstellen_alle"])}
                </div>
            </div>
            """, unsafe_allow_html=True)

        # Gesamtzusammenfassung
        total_kw = sum(ki["leistung_kw_gesamt"] for ki in kreise_info)
        st.markdown(f"""
        <div style="background:#F4F8FC;border:1.5px solid #D0E8F5;border-radius:10px;
                    padding:0.6rem 1rem;margin-top:0.4rem;font-size:0.85rem;">
            📊 <strong>Gesamt:</strong> {len(ks_list)} Kühlstellen ·
            {len(kreise_info)} Kreise ·
            {round(total_kw,2)} kW Gesamtkälteleistung
        </div>
        """, unsafe_allow_html=True)

# ============================================================
# TAB 2 – KÜHLSTELLEN
# ============================================================
with t2:

    from modules.ki_import import analysiere_excel, analysiere_bild, get_gelernte_muster

    # ----------------------------------------------------------
    # KI-IMPORT
    # ----------------------------------------------------------
    if True:  # KI-Import direkt ohne Expander

        st.markdown("""
        <div class="info-box">
        <strong>So funktioniert der KI-Import:</strong><br>
        Lade einen Kälteplan, eine Ausschreibung, eine Excel-Liste oder ein Foto hoch –
        Claude analysiert den Inhalt automatisch und legt alle erkannten Kühlstellen mit
        Temperaturen und Komponenten vor. Du prüfst, korrigierst und übernimmst mit einem Klick.
        </div>
        """, unsafe_allow_html=True)

        if not api_verfuegbar():
            st.warning("⚠️ Claude API Key fehlt – in .streamlit/secrets.toml unter [anthropic] api_key eintragen.")
        else:
            st.success("✅ Claude API aktiv – alle Formate verfügbar")

        ki1,ki2,ki3,ki4,ki5 = st.tabs(["📄 PDF","📐 DWG / DXF","📊 Excel","📷 Screenshot / Foto","✏️ Text einfügen"])

        with ki1:
            st.caption("Kälteplan, Ausschreibung oder Schema als PDF – Claude liest Text und Grafik direkt.")
            pdf_f = st.file_uploader("PDF-Datei auswählen", type=["pdf"], key="ki_pdf")
            if pdf_f and st.button("🤖 PDF analysieren", type="primary", key="btn_pdf", use_container_width=True):
                with st.spinner("Claude liest den Plan... das dauert ca. 15–30 Sekunden."):
                    st.session_state.ki_ergebnis = analysiere_pdf(pdf_f.read())

        with ki2:
            st.caption("AutoCAD-Plan als DWG oder DXF – alle Texte, Raumbezeichnungen und Layer werden extrahiert.")
            dwg_f = st.file_uploader("DWG / DXF-Datei auswählen", type=["dwg","dxf"], key="ki_dwg")
            if dwg_f and st.button("🤖 DWG analysieren", type="primary", key="btn_dwg", use_container_width=True):
                with st.spinner("Extrahiere Texte aus DWG und analysiere..."):
                    st.session_state.ki_ergebnis = analysiere_dwg(dwg_f.read(), dwg_f.name)

        with ki3:
            st.caption("Excel-Tabellen: Leistungsverzeichnis, Stückliste, Raumliste – alle Sheets werden automatisch gelesen.")
            xl_f = st.file_uploader("Excel-Datei auswählen (.xlsx, .xls)", type=["xlsx","xls"], key="ki_xl")
            if xl_f and st.button("🤖 Excel analysieren", type="primary", key="btn_xl", use_container_width=True):
                with st.spinner("Lese Excel-Tabelle und analysiere..."):
                    st.session_state.ki_ergebnis = analysiere_excel(xl_f.read(), xl_f.name)

        with ki4:
            st.caption("Screenshot, Foto oder Handskizze als Bild (PNG, JPG, JPEG, WEBP) oder PDF einer Skizze hochladen.")
            img_f = st.file_uploader("Bild / Skizze auswählen", type=["png","jpg","jpeg","webp","pdf"], key="ki_img")
            if img_f:
                if img_f.name.lower().endswith(".pdf"):
                    st.info(f"📄 PDF erkannt: {img_f.name} – wird als Dokument analysiert")
                    if st.button("🤖 Skizze/Plan analysieren", type="primary", key="btn_img", use_container_width=True):
                        with st.spinner("Claude analysiert den Plan..."):
                            img_f.seek(0)
                            st.session_state.ki_ergebnis = analysiere_pdf(img_f.read())
                else:
                    st.image(img_f, caption="Vorschau – Claude analysiert das Bild", use_column_width=True)
                    if st.button("🤖 Bild / Skizze analysieren", type="primary", key="btn_img", use_container_width=True):
                        with st.spinner("Claude analysiert das Bild..."):
                            img_f.seek(0)
                            st.session_state.ki_ergebnis = analysiere_bild(img_f.read(), img_f.name)

        with ki5:
            st.caption("Text aus E-Mail, Word-Dokument oder PDF kopieren und hier einfügen – Claude erkennt Kühlstellen, Temperaturen und Leistungen.")
            ft = st.text_area("Text hier einfügen",
                height=180,
                placeholder="Beispiel:\nKühlraum Fleisch: 0–4°C, ca. 3 kW, Verdampfer mit 2 Lüftern, EEV, Abtauheizung 230V\nTiefkühlraum: –22°C, 6 kW, Heißgasabtau, Personennotruf\nGemüselager: 4–8°C, 1,5 kW")
            if ft and st.button("🤖 Text analysieren", type="primary", key="btn_txt", use_container_width=True):
                with st.spinner("Claude analysiert den Text..."):
                    st.session_state.ki_ergebnis = analysiere_freitext(ft)

        # Gelernte Muster
        muster = get_gelernte_muster()
        if muster:
            with st.expander(f"🧠 {len(muster)} gelernte Muster aus dieser Sitzung – werden beim nächsten Import als Kontext verwendet"):
                for m in muster:
                    st.markdown(f"- **{m['name_pattern']}** → {m['temp_bereich']} · "
                                f"Raumtemp {m.get('raum_temp_soll_c','?')}°C · "
                                f"{', '.join(m.get('komponenten_erkannt',[]))}")

        # KI-Ergebnis
        if st.session_state.ki_ergebnis:
            res = st.session_state.ki_ergebnis
            st.markdown("---")
            if "fehler" in res:
                st.error(f"❌ Fehler: {res['fehler']}")
                if "tipp" in res:
                    st.info(f"💡 Tipp: {res['tipp']}")
                if "raw" in res:
                    with st.expander("Rohantwort der KI"):
                        st.text(res["raw"])
            else:
                ki_ks = res.get("kuehlstellen",[])
                ki_info2 = res.get("projekt_info",{})
                st.success(f"✅ {len(ki_ks)} Kühlstelle(n) erkannt – bitte prüfen und übernehmen!")
                if ki_info2.get("name"):
                    st.info(f"📌 Projekt erkannt: **{ki_info2['name']}** · {ki_info2.get('standort','')}")
                if ki_info2.get("bemerkung"):
                    st.markdown(f'<div class="hint-box">📝 {ki_info2["bemerkung"]}</div>', unsafe_allow_html=True)
                for h in res.get("hinweise",[]):
                    st.markdown(f'<div class="hint-box">⚠️ {h}</div>', unsafe_allow_html=True)
                # ── ERKANNTE KREISE ANZEIGEN ──
                erkannte_kreise = get_erkannte_kreise(res)
                if erkannte_kreise:
                    st.markdown("**🌡️ Erkannte Kältekreise:**")
                    kr_cols = st.columns(len(erkannte_kreise))
                    from modules.kuehlstellen import TEMP_BEREICHE
                    for ki_c, ek in enumerate(erkannte_kreise):
                        tb = ek.get("temp_bereich","NK")
                        farbe = TEMP_BEREICHE.get(tb,{}).get("farbe","#36A9E1")
                        anz = len([k for k in ki_ks if k.get("kreis") == ek["kreis_nr"]])
                        with kr_cols[ki_c]:
                            st.markdown(f"""<div style="border-left:4px solid {farbe};
                                padding:0.5rem 0.7rem;background:white;border-radius:0 8px 8px 0;
                                font-size:0.85rem;">
                                <strong style="color:{farbe};">Kreis {ek["kreis_nr"]}</strong><br>
                                {ek.get("bezeichnung","–")}<br>
                                <span style="color:#666;">{ek.get("verdampfung_c","?")}°C · {anz} Kühlst.</span>
                            </div>""", unsafe_allow_html=True)
                    st.markdown("")

                # ── INTERAKTIVE RÜCKFRAGE-KARTE ──
                st.markdown('''<div style="background:#EBF5FB;border-left:4px solid #36A9E1;
                    padding:0.8rem 1rem;border-radius:0 8px 8px 0;margin-bottom:0.8rem;">
                    <strong>📋 Bitte prüfen und ergänzen – dann übernehmen</strong><br>
                    <span style="font-size:0.83rem;color:#555;">Alle Felder sind editierbar – pro Kühlstelle nach Übernahme weiter anpassbar</span>
                </div>''', unsafe_allow_html=True)

                # Maschinenstandort wählen
                ms_opts = st.session_state.get("maschinenstandorte",
                    [{"id":"ms1","anlage_typ":"Verbundanlage","standort_maschine":"Maschinenraum EG"}])
                ms_labels = [f"{m.get('anlage_typ',m.get('name','?'))} – {m.get('standort_maschine',m.get('etage',''))}" for m in ms_opts]
                ki_ms_idx = st.selectbox("🏭 Maschinenstandort / Aggregatstandort",
                    range(len(ms_labels)), format_func=lambda i: ms_labels[i],
                    key="ki_ms_sel",
                    help="Von welchem Standort werden die Leitungslängen gemessen?")

                # Globaler Lieferumfang
                rk_c1, rk_c2 = st.columns(2)
                with rk_c1:
                    ki_lf_alle = st.radio("🔧 Lieferumfang (alle Kühlstellen)",
                        ["Wie erkannt", "Alle Direkt", "Alle Extern"],
                        horizontal=True, key="ki_lf_preset",
                        help="Schnell-Einstellung – pro Kühlstelle nach Übernahme editierbar")

                st.markdown("---")
                st.markdown("**📏 Leitungslängen & Lieferumfang pro Kühlstelle:**")
                st.caption("Leitungslänge = vom Maschinenstandort zur Kühlstelle. 0 = direkt nebenan.")

                # Rückfrage-Tabelle pro Kühlstelle
                ki_laengen = {}
                ki_lf_einzeln = {}
                ki_lieferumfang_check = {}

                for k_idx, k in enumerate(ki_ks):
                    rk_row = st.columns([3,1,2,2,2])
                    with rk_row[0]:
                        pos = f"Pos. {k['pos_nr']} – " if k.get("pos_nr") else ""
                        st.markdown(f"**{pos}{k.get('name','?')}**")
                        st.caption(f"Kreis {k.get('kreis','?')} · {k.get('temp_bereich','?')} · {k.get('raum_temp_soll_c','?')}°C · {k.get('kaelteleistung_kw','?')} kW")
                    with rk_row[1]:
                        ki_laengen[k_idx] = st.number_input(
                            "Kreis", 1, 20, int(k.get("kreis",1)),
                            key=f"ki_kreis_{k_idx}",
                            label_visibility="collapsed")
                        st.caption("Kreis Nr.")
                    with rk_row[2]:
                        ki_laengen[k_idx] = st.number_input(
                            "Länge [m]", 0, 500,
                            int(k.get("leitungslaenge_m", 20)),
                            key=f"ki_len_{k_idx}",
                            label_visibility="collapsed",
                            help="Leitungslänge vom Maschinenstandort zu dieser Kühlstelle")
                        st.caption("→ Maschinenstandort [m]")
                    with rk_row[3]:
                        lf_default = k.get("lieferumfang","direkt")
                        if ki_lf_alle == "Alle Direkt": lf_default = "direkt"
                        elif ki_lf_alle == "Alle Extern": lf_default = "extern"
                        ki_lf_einzeln[k_idx] = st.radio(
                            "Lieferumfang",
                            ["direkt","extern"],
                            format_func=lambda x: "🔧 Direkt" if x=="direkt" else "🔄 Extern",
                            index=["direkt","extern"].index(lf_default),
                            key=f"ki_lf_{k_idx}",
                            horizontal=True)
                    with rk_row[4]:
                        st.markdown(f"Komp.: {len(k.get('komponenten_erkannt',[]))}")
                        if k.get("komponenten_erkannt"):
                            st.caption(", ".join(k["komponenten_erkannt"][:2]) +
                                      ("..." if len(k.get("komponenten_erkannt",[])) > 2 else ""))

                st.markdown("---")
                rk_btn1, rk_btn2 = st.columns(2)
                with rk_btn1:
                    if st.button("✅ Übernehmen & Zusammenführen", type="primary", use_container_width=True):
                        neue_liste = ki_ergebnis_zu_kuehlstellen(res)
                        # Leitungslängen + Lieferumfang aus Rückfrage-Karte anwenden
                        for k_i, ks_item in enumerate(neue_liste):
                            ks_item["leitungslaenge_m"] = ki_laengen.get(k_i, 20)
                            ks_item["lieferumfang"] = ki_lf_einzeln.get(k_i, "direkt")
                            if ms_opts and ki_ms_idx < len(ms_opts):
                                ms_sel = ms_opts[ki_ms_idx]
                                ks_item["maschinenstandort"] = (
                                    ms_sel.get("anlage_typ","") + " – " +
                                    ms_sel.get("standort_maschine",""))
                        # Merge mit bestehenden (Duplikate erkennen + aktualisieren)
                        bestehende = st.session_state.kuehlstellen
                        merged, neu_c, upd_c = merge_kuehlstellen(bestehende, neue_liste)
                        st.session_state.kuehlstellen = merged
                        st.session_state.ki_ergebnis = None
                        if neu_c > 0 and upd_c > 0:
                            st.success(f"✅ {neu_c} neue Kühlstellen hinzugefügt · {upd_c} bestehende aktualisiert")
                        elif neu_c > 0:
                            st.success(f"✅ {neu_c} neue Kühlstellen hinzugefügt")
                        elif upd_c > 0:
                            st.success(f"✅ {upd_c} bestehende Kühlstellen aktualisiert (keine Duplikate)")
                        st.rerun()
                with rk_btn2:
                    if st.button("🔄 Verwerfen", use_container_width=True):
                        st.session_state.ki_ergebnis = None; st.rerun()

    st.markdown("---")

    # ----------------------------------------------------------
    # DATENBLATT-IMPORT
    # ----------------------------------------------------------
    with st.expander("📋 Gerätedaten / Datenblätter hochladen (Verdampfer, Außenunits)", expanded=False):
        st.caption("Lade Datenblätter hoch – die KI liest alle technischen Daten und speichert sie in der Gerätedatenbank.")
        db_kl = st.session_state.db
        dbl_tab1, dbl_tab2, dbl_tab3 = st.tabs(["📄 PDF Datenblatt", "📷 Bild / Screenshot", "📊 Excel Stückliste"])

        with dbl_tab1:
            dbl_pdf = st.file_uploader("Datenblatt als PDF", type=["pdf"], key="dbl_pdf")
            if dbl_pdf and st.button("🤖 Datenblatt analysieren", type="primary", key="btn_dbl_pdf"):
                with st.spinner("KI liest Datenblatt..."):
                    dbl_pdf.seek(0)
                    dbl_res = analysiere_datenblatt_pdf(dbl_pdf.read())
                    st.session_state["dbl_result"] = dbl_res

        with dbl_tab2:
            dbl_img = st.file_uploader("Datenblatt als Bild", type=["png","jpg","jpeg","webp"], key="dbl_img")
            if dbl_img and st.button("🤖 Bild analysieren", type="primary", key="btn_dbl_img"):
                with st.spinner("KI liest Gerätedaten..."):
                    dbl_img.seek(0)
                    dbl_res = analysiere_datenblatt_bild(dbl_img.read(), dbl_img.name)
                    st.session_state["dbl_result"] = dbl_res

        with dbl_tab3:
            dbl_xl = st.file_uploader("Geräteliste als Excel", type=["xlsx","xls"], key="dbl_xl")
            if dbl_xl and st.button("🤖 Excel analysieren", type="primary", key="btn_dbl_xl"):
                with st.spinner("KI liest Geräteliste..."):
                    dbl_xl.seek(0)
                    dbl_res = analysiere_datenblatt_excel(dbl_xl.read())
                    st.session_state["dbl_result"] = dbl_res

        # Ergebnis anzeigen + speichern
        if st.session_state.get("dbl_result"):
            dbl_res = st.session_state["dbl_result"]
            if "fehler" in dbl_res:
                st.error(f"❌ {dbl_res['fehler']}")
            else:
                # Einzelnes Gerät oder Liste
                geraete_liste = dbl_res.get("geraete", [dbl_res]) if "geraete" in dbl_res else [dbl_res]
                for g_idx, g in enumerate(geraete_liste):
                    st.markdown(f"**Erkanntes Gerät {g_idx+1}:** {g.get('hersteller','')} {g.get('modell','')} – {g.get('typ','')} – {g.get('kaelteleistung_kw','?')} kW")
                    with st.expander("Details anzeigen", expanded=len(geraete_liste)==1):
                        gc1, gc2, gc3 = st.columns(3)
                        with gc1:
                            st.metric("Kälteleistung", f"{g.get('kaelteleistung_kw','?')} kW")
                            st.metric("Lüfter", f"{g.get('anzahl_luefter','?')} × {g.get('leistung_luefter_w','?')} W")
                            st.metric("Spannung", g.get("spannung_luefter", g.get("anschluss_spannung","?")))
                        with gc2:
                            st.metric("Raumtemp.", f"{g.get('raum_temp_c','?')}°C")
                            st.metric("Verdampfung", f"{g.get('verdampfung_c','?')}°C")
                            st.metric("Abtau", g.get("abtau_typ","?"))
                        with gc3:
                            st.metric("Kältemittel", g.get("kaeltemittel","?"))
                            st.metric("Bus", g.get("bus_schnittstelle","–") or "–")
                            st.metric("Typ", g.get("typ","?"))

                if st.button("💾 In Gerätedatenbank speichern", type="primary", key="dbl_save"):
                    for g in geraete_liste:
                        db_kl = add_geraet(db_kl, g.copy(), get_display_name())
                    speichere_datenbank(db_kl)
                    st.session_state.db = db_kl
                    st.session_state["dbl_result"] = None
                    st.success(f"✅ {len(geraete_liste)} Gerät(e) gespeichert!")
                    st.rerun()

    # Gerätedatenbank-Übersicht
    db_kl = st.session_state.db
    alle_geraete = get_geraete(db_kl)
    if alle_geraete:
        with st.expander(f"🗄️ Gerätedatenbank – {len(alle_geraete)} Geräte gespeichert", expanded=False):
            for g in alle_geraete:
                gc1, gc2, gc3 = st.columns([3,3,1])
                with gc1:
                    st.markdown(f"**{g.get('id','')}** {g.get('hersteller','')} {g.get('modell','')}")
                    st.caption(f"{g.get('typ','')} · {g.get('kaelteleistung_kw','?')} kW · {g.get('kaeltemittel','')}")
                with gc2:
                    st.caption(f"Lüfter: {g.get('anzahl_luefter','?')}× {g.get('leistung_luefter_w','?')}W · Abtau: {g.get('abtau_typ','?')}")
                with gc3:
                    if st.button("🗑️", key=f"del_g_{g['id']}"):
                        db_kl = delete_geraet(db_kl, g["id"])
                        speichere_datenbank(db_kl)
                        st.session_state.db = db_kl
                        st.rerun()

    st.markdown("---")

    # ----------------------------------------------------------
    # KÜHLSTELLEN LISTE + FORMULAR
    # ----------------------------------------------------------
    cl, cr = st.columns([1, 2], gap="large")

    # LISTE LINKS
    with cl:
        st.markdown('<div class="sec">❄️ Kühlstellen im Projekt</div>', unsafe_allow_html=True)
        st.caption("Klicke eine Kühlstelle an um sie rechts zu bearbeiten.")

        cn1,cn2 = st.columns(2)
        with cn1:
            if st.button("➕ Neue Kühlstelle", use_container_width=True, type="primary"):
                nr = len(st.session_state.kuehlstellen)+1
                ks_neu = neue_kuehlstelle(nr)
                # Globale Defaults anwenden
                ks_neu["komponenten"]["tuer_kontakt"] = {
                    "aktiv": st.session_state.get("global_tuer_kontakt", True),
                    "parameter": {"anzahl_tueren": st.session_state.get("global_tueren", 1),
                                  "typ": "NC (Schließer)", "tuerlicht_schaltung": True},
                    "ki_erkannt": False
                }
                ks_neu["komponenten"]["innenbeleuchtung"] = {
                    "aktiv": st.session_state.get("global_licht", True),
                    "parameter": {"licht_typ": "LED feuchtraumgeeignet IP65",
                                  "spannung": "230V 1-phasig",
                                  "schaltung": "Türkontakt (automatisch)", "leistung_w": 30},
                    "ki_erkannt": False
                }
                ks_neu["komponenten"]["bewegungsmelder"] = {
                    "aktiv": st.session_state.get("global_bewegungsmelder", True),
                    "parameter": {"typ_melder": "PIR (passiv infrarot, Standard)",
                                  "funktion": "Lichtschaltung", "spannung": "230V 1-phasig"},
                    "ki_erkannt": False
                }
                ks_neu["komponenten"]["haccp_aufzeichnung"] = {
                    "aktiv": st.session_state.get("global_haccp", True),
                    "parameter": {
                        "anzahl_fuehler": 1,
                        "fuehler_typ": "NTC (Standard)",
                        "aufzeichnung_system": "Regler-intern (Datalogging im Regler)",
                        "alarm_grenzwert_oben": 8,
                        "alarm_grenzwert_unten": -1,
                        "alarm_verzoegerung_min": 30,
                        "haccp_klasse": "Standard (EU-VO 37/2005)"
                    },
                    "ki_erkannt": False
                }
                st.session_state.kuehlstellen.append(ks_neu)
                st.session_state.aktive_ks_idx = len(st.session_state.kuehlstellen)-1; st.rerun()
        with cn2:
            if st.session_state.kuehlstellen:
                if st.button("📋 Duplizieren", use_container_width=True,
                             help="Kopiert die aktuell ausgewählte Kühlstelle mit allen Einstellungen"):
                    idx = st.session_state.aktive_ks_idx
                    nr = len(st.session_state.kuehlstellen)+1
                    st.session_state.kuehlstellen.append(
                        kopiere_kuehlstelle(st.session_state.kuehlstellen[idx], nr))
                    st.session_state.aktive_ks_idx = len(st.session_state.kuehlstellen)-1; st.rerun()

        st.markdown("<br>", unsafe_allow_html=True)

        for i, ks in enumerate(st.session_state.kuehlstellen):
            ti = get_temp_bereich_info(ks.get("temp_bereich","NK"))
            lf_i = "🔧" if ks.get("lieferumfang")=="direkt" else "🔄"
            ki_flag = "🤖 " if any(v.get("ki_erkannt") for v in ks.get("komponenten",{}).values()) else ""
            is_act = (i == st.session_state.aktive_ks_idx)
            aktive_komps = len([v for v in ks.get("komponenten",{}).values() if v.get("aktiv")])

            # Infos aufbauen
            e_nr   = ks.get("e_nr", "")
            pos_nr = ks.get("pos_nr", "")
            gk_pos = ks.get("lv_pos", "")
            kreis  = ks.get("kreis", "")
            laenge = ks.get("leitungslaenge_m", "")

            zeile1 = f"{ki_flag}{lf_i} Kreis {kreis}  ·  {e_nr}  ·  Pos.{pos_nr}  ·  {gk_pos}"
            zeile2 = f"{ks['name']}  ·  {ti['label']}  ·  {ks.get('verdampfung_custom_c', ti.get('verdampfung_c','?'))}°C  ·  {ks.get('kaelteleistung_kw','?')} kW  ·  {laenge}m  ·  {aktive_komps} Komp."

            ca,cb = st.columns([4,1])
            with ca:
                if st.button(
                    f"{zeile1}\n{zeile2}",
                    key=f"ksb_{i}", use_container_width=True,
                    type="primary" if is_act else "secondary"):
                    st.session_state.aktive_ks_idx = i; st.rerun()
            with cb:
                if st.button("🗑️", key=f"dks_{i}", help="Kühlstelle löschen"):
                    st.session_state.kuehlstellen.pop(i)
                    st.session_state.aktive_ks_idx = max(0, i-1); st.rerun()

    # FORMULAR RECHTS
    with cr:
        if not st.session_state.kuehlstellen:
            st.markdown("""
            <div class="info-box">
            <strong>Noch keine Kühlstellen vorhanden.</strong><br>
            ➕ Neue Kühlstelle anlegen – oder oben den KI-Import verwenden um Kühlstellen
            automatisch aus einem Plan oder einer Ausschreibung zu erkennen.
            </div>
            """, unsafe_allow_html=True)
        else:
            idx = min(st.session_state.aktive_ks_idx, len(st.session_state.kuehlstellen)-1)
            ks = st.session_state.kuehlstellen[idx]

            # ── SEKTION 1: GRUNDDATEN ──
            st.markdown(f'<div class="sec">📋 Grunddaten – {ks["name"]}</div>', unsafe_allow_html=True)
            st.caption("Allgemeine Informationen zur Kühlstelle – Name, Lage, Zuständigkeit und technische Basisdaten.")

            e1,e2 = st.columns(2)
            with e1:
                ks["name"] = st.text_input(
                    "Bezeichnung der Kühlstelle *",
                    value=ks["name"], key=f"kn_{idx}",
                    placeholder="z.B. Kühlraum Fleisch EG, TK-Lager 1, Gemüsekühlung",
                    help="Eindeutige Bezeichnung – erscheint in allen Auswertungen und der Kabelliste")
                ks["beschreibung"] = st.text_input(
                    "Kurzbeschreibung / Lage",
                    value=ks.get("beschreibung",""), key=f"kd_{idx}",
                    placeholder="z.B. EG links neben Anlieferung, 2. OG Produktionshalle",
                    help="Hilft bei der Orientierung im Plan und auf der Baustelle")
                ks["lieferumfang"] = st.radio(
                    "Lieferumfang",
                    ["direkt","extern"],
                    format_func=lambda x: "🔧 Direkt – unser Gewerk (wir liefern & montieren)" if x=="direkt"
                                         else "🔄 Extern – Fremdgewerk (Elektriker / Beistellung Kunde)",
                    index=["direkt","extern"].index(ks.get("lieferumfang","direkt")),
                    horizontal=False, key=f"klf_{idx}",
                    help="Direkt = Kabel und Montage in unserem Angebot enthalten. Extern = anderes Gewerk, wir dokumentieren nur.")

            with e2:
                ks["kreis"] = st.number_input(
                    "Kältekreis Nr.",
                    min_value=1, max_value=20, value=int(ks.get("kreis",1)), key=f"kkr_{idx}",
                    help="Kühlstellen mit gleicher Verdampfungstemperatur gehören zum selben Kreis – wichtig für Verbundanlage/Maschinenauswahl")
                ks["kaelteleistung_kw"] = st.number_input(
                    "Kälteleistung [kW]",
                    min_value=0.1, max_value=500.0,
                    value=float(ks.get("kaelteleistung_kw") or 2.0),
                    step=0.1, key=f"kkw_{idx}",
                    help="Nennkälteleistung des Verdampfers laut Datenblatt oder Auslegung")

            # ── SEKTION 2: TEMPERATUREN ──
            # Gerät zuweisen
            alle_geraete_ks = get_geraete(st.session_state.db)
            if alle_geraete_ks:
                st.markdown('<div class="sec">🔧 Gerät auswählen (aus Datenbank)</div>', unsafe_allow_html=True)
                vd_geraete = [g for g in alle_geraete_ks if g.get("typ") == "Verdampfer"]
                au_geraete = [g for g in alle_geraete_ks if "Außenunit" in g.get("typ","") or "Verflüssiger" in g.get("typ","") or "Kompakt" in g.get("typ","")]
                gv1, gv2 = st.columns(2)
                with gv1:
                    if vd_geraete:
                        vd_opts = ["– Kein Verdampfer gewählt –"] + [f"{g['id']} {g['hersteller']} {g['modell']} ({g.get('kaelteleistung_kw','?')} kW)" for g in vd_geraete]
                        vd_sel = st.selectbox("❄️ Verdampfer", vd_opts, key=f"vd_sel_{idx}",
                            index=next((i+1 for i,g in enumerate(vd_geraete) if g["id"]==ks.get("geraet_verdampfer_id","")), 0))
                        if vd_sel != "– Kein Verdampfer gewählt –" and st.button("✅ Verdampfer übernehmen", key=f"vd_ueb_{idx}"):
                            vd_id = vd_sel.split(" ")[0]
                            g_sel = next((g for g in vd_geraete if g["id"]==vd_id), None)
                            if g_sel:
                                ks = geraet_zu_kuehlstelle(g_sel, ks)
                                st.session_state.kuehlstellen[idx] = ks
                                st.success(f"✅ {g_sel['hersteller']} {g_sel['modell']} übernommen!")
                                st.rerun()
                    if ks.get("geraet_verdampfer_modell"):
                        st.caption(f"✅ Aktuell: {ks['geraet_verdampfer_modell']}")
                with gv2:
                    if au_geraete:
                        au_opts = ["– Keine Außenunit gewählt –"] + [f"{g['id']} {g['hersteller']} {g['modell']} ({g.get('kaelteleistung_kw','?')} kW)" for g in au_geraete]
                        au_sel = st.selectbox("🏭 Außenunit / Verflüssiger", au_opts, key=f"au_sel_{idx}",
                            index=next((i+1 for i,g in enumerate(au_geraete) if g["id"]==ks.get("geraet_aussenunit_id","")), 0))
                        if au_sel != "– Keine Außenunit gewählt –" and st.button("✅ Außenunit übernehmen", key=f"au_ueb_{idx}"):
                            au_id = au_sel.split(" ")[0]
                            g_sel = next((g for g in au_geraete if g["id"]==au_id), None)
                            if g_sel:
                                ks = geraet_zu_kuehlstelle(g_sel, ks)
                                st.session_state.kuehlstellen[idx] = ks
                                st.success(f"✅ {g_sel['hersteller']} {g_sel['modell']} übernommen!")
                                st.rerun()
                    if ks.get("geraet_aussenunit_modell"):
                        st.caption(f"✅ Aktuell: {ks['geraet_aussenunit_modell']}")

            st.markdown('<div class="sec">🌡️ Temperaturen & Verdampfungskreis</div>', unsafe_allow_html=True)
            st.caption("Raumtemperatur und Verdampfungstemperatur bestimmen die Kreiszuordnung und später die Maschinenauswahl.")

            topt = list(TEMP_BEREICHE.keys())
            t1c,t2c,t3c = st.columns(3)
            with t1c:
                ks["temp_bereich"] = st.selectbox(
                    "Temperaturbereich",
                    topt,
                    format_func=lambda x: f"{TEMP_BEREICHE[x]['label']}",
                    index=topt.index(ks.get("temp_bereich","NK")),
                    key=f"ktb_{idx}",
                    help="Vordefinierte Bereiche mit typischen Temperaturen – bei Sonderfällen 'Benutzerdefiniert' wählen")
            ti = get_temp_bereich_info(ks["temp_bereich"])
            with t2c:
                rd = ti["raum_temp_c"][0] or 0
                ks["raum_temp_soll_c"] = st.number_input(
                    "Raumtemperatur Soll [°C]",
                    min_value=-50.0, max_value=50.0,
                    value=float(ks.get("raum_temp_soll_c") or rd),
                    step=0.5, key=f"krt_{idx}",
                    help="Gewünschte Lagertemperatur im Raum – aus Ausschreibung oder Lebensmittelvorschriften")
            with t3c:
                vd = ti.get("verdampfung_c") or -8
                ks["verdampfung_custom_c"] = st.number_input(
                    "Verdampfungstemperatur [°C]",
                    min_value=-60.0, max_value=20.0,
                    value=float(ks.get("verdampfung_custom_c") or vd),
                    step=0.5, key=f"kvt_{idx}",
                    help="Verdampfungstemperatur des Kältemittels – bestimmt den Kältekreis. NK: ca. -8°C / TK: ca. -31 bis -33°C")

            if ti.get('beispiele'):
                st.caption(f"💡 {ti['label']}: Raumtemp. {ti['raum_temp_c'][0]}°C–{ti['raum_temp_c'][1]}°C · Verdampfung ca. {ti.get('verdampfung_c','?')}°C · {ti['beispiele']}")

            # Regler aus Datenbank laden für Dropdown
            _regler_optionen_db = get_regler_optionen(st.session_state.db)

            # ── SEKTION 3: KOMPONENTEN ──
            st.markdown('<div class="sec">🔧 Komponenten & Anschlüsse</div>', unsafe_allow_html=True)
            st.caption("Wähle alle Komponenten die an dieser Kühlstelle vorhanden sind. "
                       "* = Pflichtkomponente · 🤖 = von KI erkannt · Vorschläge sind editierbar.")

            if "komponenten" not in ks:
                ks["komponenten"] = {}

            for gruppe in KOMPONENTEN_GRUPPEN:
                gkomps = {k:v for k,v in KOMPONENTEN.items() if v.get("gruppe")==gruppe}
                if not gkomps: continue

                _grp_color = "#36A9E1"
                if gruppe == "HACCP": _grp_color = "#27AE60"
                elif gruppe == "Sicherheit": _grp_color = "#E74C3C"
                st.markdown(f"""
                <div style="background:#F4F8FC;border-left:3px solid {_grp_color};
                            padding:4px 10px;border-radius:0 6px 6px 0;
                            margin:0.8rem 0 0.4rem 0;font-weight:600;font-size:0.88rem;">
                    {gruppe}
                </div>""", unsafe_allow_html=True)

                # HACCP: Hinweis bei Extern-Lieferumfang (Kühlmöbel/Gastro)
                if gruppe == "HACCP" and ks.get("lieferumfang") == "extern":
                    st.markdown('''<div style="background:#FEF9E7;border-left:3px solid #F39C12;
                        padding:6px 10px;border-radius:0 6px 6px 0;font-size:0.82rem;margin-bottom:0.4rem;">
                        🔄 <strong>Extern-Lieferumfang:</strong> Bei Kühlmöbeln / Gastro-Geräten ist der Regler
                        meist im Gerät integriert. HACCP-Fühler und Busleitung zum Regler (RS485) sind dann
                        theoretisch die einzigen Leitungen die noch extern verlegt werden müssen.
                    </div>''', unsafe_allow_html=True)

                for kk, ki_info in gkomps.items():
                    if kk not in ks["komponenten"]:
                        ks["komponenten"][kk] = {
                            "aktiv": ki_info.get("pflicht",False),
                            "parameter": {pk:pv.get("vorschlag") for pk,pv in ki_info.get("parameter",{}).items()},
                            "ki_erkannt": False
                        }
                    kst = ks["komponenten"][kk]
                    ki_fl = " 🤖" if kst.get("ki_erkannt") else ""
                    pfl_fl = " *" if ki_info.get("pflicht") else ""

                    ca2,cb2 = st.columns([1,3])
                    with ca2:
                        aktiv = st.checkbox(
                            f"{ki_info['icon']} {ki_info['label']}{pfl_fl}{ki_fl}",
                            value=kst["aktiv"],
                            key=f"kmp_{idx}_{kk}",
                            help=f"Gruppe: {ki_info['gruppe']}")
                        kst["aktiv"] = aktiv

                    if aktiv:
                        with cb2:
                                params = kst.get("parameter",{})
                                for pk, pd_def in ki_info.get("parameter",{}).items():
                                    ptyp = pd_def.get("typ")
                                    pkey2 = f"pr_{idx}_{kk}_{pk}"
                                    plabel = pd_def.get("label", pk)
                                    phinweis = pd_def.get("hinweis","")
                                    pvorschlag = pd_def.get("vorschlag")

                                    if ptyp == "auswahl":
                                        opts = pd_def.get("optionen",[])
                                        cur = params.get(pk, pvorschlag)
                                        if cur not in opts: cur = opts[0]
                                        params[pk] = st.selectbox(
                                            plabel, opts,
                                            index=opts.index(cur),
                                            key=pkey2,
                                            help=phinweis or f"Vorschlag: {pvorschlag}")
                                        if phinweis:
                                            st.caption(f"💡 {phinweis}")
                                        # EEV Motorkabel Warnung
                                        if kk == "eev" and pk == "laenge_motorkabel_m":
                                            _mot_len = params.get("laenge_motorkabel_m", 3)
                                            _eev_t = params.get("eev_typ","")
                                            if "Schrittmotor" in _eev_t and float(_mot_len) > 8:
                                                st.error("🚨 FEHLER: Motorkabel > 8m! Treiber näher am Ventil montieren!")
                                            elif "Schrittmotor" in _eev_t and float(_mot_len) > 5:
                                                st.warning("⚠️ Motorkabel >5m – Grenzbereich! Max. 8m laut Hersteller.")
                                            elif "Schrittmotor" in _eev_t:
                                                st.success(f"✅ Motorkabel {_mot_len}m – OK (max. 8m)")
                                        # Fremdsteuerung Hinweis
                                        if kk == "schaltkasten" and pk == "steuerung_typ":
                                            _st_val = params.get("steuerung_typ","")
                                            if "Bus-kompatibel" in _st_val:
                                                st.markdown('''<div style="background:#E8F8F0;border-left:3px solid #27AE60;padding:6px 10px;border-radius:0 6px 6px 0;font-size:0.82rem;">
                                                ✅ <strong>Bus-kompatibel:</strong> Nur Busleitung (RS485) und ggf. HACCP-Fühler verlegen. Kein eigener Regler nötig.
                                                </div>''', unsafe_allow_html=True)
                                            elif "Bus NICHT kompatibel" in _st_val:
                                                st.markdown('''<div style="background:#FEF9E7;border-left:3px solid #E67E22;padding:6px 10px;border-radius:0 6px 6px 0;font-size:0.82rem;">
                                                ⚠️ <strong>Bus NICHT kompatibel:</strong> Kein Datenbus möglich. Nur potenzialfreie Kontakte (Störmeldung, Abtausignal) verdrahtbar. Kein Monitoring via RS485.
                                                </div>''', unsafe_allow_html=True)

                                    elif ptyp == "zahl":
                                        params[pk] = st.number_input(
                                            plabel,
                                            min_value=float(pd_def.get("min",0)),
                                            max_value=float(pd_def.get("max",9999)),
                                            value=float(params.get(pk, pvorschlag) or pvorschlag),
                                            key=pkey2,
                                            help=phinweis or f"Typischer Wert: {pvorschlag}")

                                    elif ptyp == "bool":
                                        params[pk] = st.checkbox(
                                            plabel,
                                            value=bool(params.get(pk, pvorschlag)),
                                            key=pkey2,
                                            help=phinweis)

                                kst["parameter"] = params

            # ── SEKTION 4: STANDORT & LEITUNGEN ──
            st.markdown('<div class="sec">📍 Standort & Leitungslängen</div>', unsafe_allow_html=True)
            st.caption("Standortangaben und Kabellängen für die Kabelberechnung und Dokumentation.")
            loc1, loc2 = st.columns(2)
            with loc1:
                ks["standort_raum"] = st.text_input(
                    "Raumbezeichnung / Lage",
                    value=ks.get("standort_raum",""),
                    key=f"slr_{idx}",
                    placeholder="z.B. Kühlraum EG links, TK-Lager 2.OG",
                    help="Genaue Lage im Gebäude – für Montageplan und Dokumentation")
                ks["standort_etage"] = st.text_input(
                    "Etage / Bereich",
                    value=ks.get("standort_etage",""),
                    key=f"sle_{idx}",
                    placeholder="z.B. EG, 1.OG, Keller, Außen",
                    help="Stockwerk oder Bereich")
            with loc2:
                # SK-Status für diese Kühlstelle bestimmen
                _sk_override = ks.get("zentraler_sk_override", None)
                _sk_zentral = _sk_override if _sk_override is not None else st.session_state.get("globaler_sk", False)
                _regler_an_zelle = not _sk_zentral

                # Kabel Kühlstelle → Regler/SK
                if _regler_an_zelle:
                    # Regler direkt an Zelle – kein langes Kabel nötig
                    ks["leitungslaenge_m"] = st.number_input(
                        "Kabel Kühlstelle → Regler [m]",
                        0, 500, int(ks.get("leitungslaenge_m", 2)),
                        key=f"kll_{idx}",
                        help="Regler sitzt direkt an der Zelle → kurzes Kabel (0–5m typisch)")
                    st.caption("ℹ️ Regler an Zelle – kurze Leitungen")
                else:
                    ks["leitungslaenge_m"] = st.number_input(
                        "Kabel Kühlstelle → Zentraler SK [m]",
                        0, 500, int(ks.get("leitungslaenge_m", 20)),
                        key=f"kll_{idx}",
                        help="Leitungslänge vom zentralen Schaltschrank zur Kühlstelle – beeinflusst Kabelquerschnitt")

                ks["laenge_aussenteil_m"] = st.number_input(
                    "Kabel → Außenteil / Verflüssiger [m]",
                    0, 500, int(ks.get("laenge_aussenteil_m", 15)),
                    key=f"lat_{idx}",
                    help="0 = kein Außenteil (z.B. Kompaktaggregat). Sonst Länge zum Außenteil/Verflüssiger")
                ks["laenge_router_m"] = st.number_input(
                    "Kabel → Router / Netzwerk [m]",
                    0, 500, int(ks.get("laenge_router_m", 20)),
                    key=f"lro_{idx}",
                    help="0 = kein Netzwerk/Monitoring. Sonst Länge LAN/RS485 zum Router oder Switch")

                # Zentraler SK: immer anzeigen aber nur wenn SK aktiv relevant
                if _sk_zentral:
                    ks["laenge_schaltschrank_m"] = st.number_input(
                        "Kabel → Zentraler Schaltschrank [m]",
                        0, 500, int(ks.get("laenge_schaltschrank_m", 10)),
                        key=f"lsk_{idx}",
                        help="Leitungslänge zur Sammelschiene des zentralen SK – für HACCP-Module, GWA-Versorgung, Verteiler")

            # Schaltschrank Override
            st.markdown('<div class="sec">🔌 Schaltschrank – Einstellung für diese Kühlstelle</div>', unsafe_allow_html=True)
            globaler_sk = st.session_state.get("globaler_sk", False)
            sk_override = ks.get("zentraler_sk_override", None)
            sk_optionen = ["Globale Einstellung übernehmen", "Einzelregler an der Zelle", "Zentraler Schaltschrank"]
            if sk_override is None:
                sk_idx = 0
            elif sk_override == False:
                sk_idx = 1
            else:
                sk_idx = 2
            sk_wahl = st.radio(
                "Schaltschrank für diese Kühlstelle",
                sk_optionen,
                index=sk_idx,
                horizontal=True,
                key=f"sko_{idx}",
                help="Überschreibt die globale Einstellung aus Tab 1")
            if sk_wahl == "Globale Einstellung übernehmen":
                ks["zentraler_sk_override"] = None
                sk_aktiv = globaler_sk
            elif sk_wahl == "Einzelregler an der Zelle":
                ks["zentraler_sk_override"] = False
                sk_aktiv = False
            else:
                ks["zentraler_sk_override"] = True
                sk_aktiv = True

            if sk_aktiv:
                st.caption(f"🏭 Zentraler Schaltschrank: {st.session_state.get('globaler_sk_bezeichnung','SK-01')} · "
                           f"Bedienteil / Display trotzdem an Zelle möglich (Regler-Einstellung oben)")
            else:
                st.caption("🔌 Einzelregler direkt an der Kühlzelle montiert")

            # ── SEKTION 5: NOTIZEN ──
            st.markdown('<div class="sec">📝 Notizen</div>', unsafe_allow_html=True)
            ks["notizen"] = st.text_area(
                "Notizen zur Kühlstelle",
                value=ks.get("notizen",""),
                key=f"kno_{idx}",
                height=80,
                placeholder="z.B. Heißgasabtau laut Plan, 2 Lüfter erkannt, Türkontakt vorhanden, besondere Anforderungen...",
                help="Freies Notizfeld – erscheint in der Dokumentation")

            # Zurückschreiben
            st.session_state.kuehlstellen[idx] = ks

# ============================================================
# TAB 3 – STEUERUNG
# ============================================================
with t3:
    st.markdown('<div class="sec">🎛️ Steuerungssystem</div>', unsafe_allow_html=True)
    cs1,cs2 = st.columns([1,2], gap="large")
    with cs1:
        for skey, sinfo in STEUERUNGSSYSTEME.items():
            is_sel = (st.session_state.steuerung == skey)
            if st.button(f"{sinfo['icon']} {sinfo['label']}", key=f"st_{skey}",
                         use_container_width=True, type="primary" if is_sel else "secondary"):
                st.session_state.steuerung = skey; st.rerun()
    with cs2:
        if st.session_state.steuerung:
            si = STEUERUNGSSYSTEME[st.session_state.steuerung]
            st.markdown(f'''<div class="ks-card">
                <div style="font-size:1.6rem;margin-bottom:4px;">{si["icon"]} <strong style="font-size:1.1rem;">{si["label"]}</strong></div>
                <div style="color:#555;font-size:0.88rem;margin-bottom:0.8rem;">{si["beschreibung"]}</div>
                <div style="display:flex;gap:1.5rem;font-size:0.85rem;">
                    <span>🔌 <strong>Bus/Protokoll:</strong> {si.get("bus") or "Kein Bus"}</span>
                    <span>🏗️ <strong>Typisch für:</strong> {si["typisch_fuer"]}</span>
                </div>
            </div>''', unsafe_allow_html=True)

            # Zubehör anzeigen
            zub = si.get("zubehoer", si.get("zubehör", []))
            if zub:
                st.markdown("**📦 Zubehör & Systemkomponenten:**")
                for z in zub:
                    st.markdown(f"&nbsp;&nbsp;• {z}")

            # Kabeltypen
            rows = []
            for k in si.get("kabel",[]):
                for cl_d in KABEL_MATRIX.values():
                    for kt in cl_d["kabel_typen"]:
                        if kt["kuerzel"]==k:
                            rows.append({"Kürzel":kt["kuerzel"],"Bezeichnung":kt["bezeichnung"],"Typ":kt["typ"],"Norm":kt["norm"]})
            if rows:
                st.markdown("**🔌 Benötigte Kabeltypen:**")
                st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)
        else:
            st.markdown('<div class="info-box">⬅️ Steuerungssystem auswählen.</div>', unsafe_allow_html=True)

    st.markdown("---")
    st.markdown('<div class="sec">🔗 Kreise & Schaltungszuordnung</div>', unsafe_allow_html=True)
    ks_list2 = st.session_state.kuehlstellen
    if not ks_list2:
        st.info("Zuerst Kühlstellen in Tab 2 erfassen.")
    else:
        kreise_info2 = kreis_zusammenfassung(ks_list2)

        for ki_idx, ki in enumerate(kreise_info2):
            farbe = ki["farbe"]
            exp_label = (f"Kreis {ki['kreis_nr']} – {ki['temp_label']}  "
                        f"·  {ki['verdampfung_c']}°C  "
                        f"·  {ki['anzahl_gesamt']} Kühlstellen  "
                        f"·  {ki['leistung_kw_gesamt']} kW")

            with st.expander(exp_label, expanded=True):
                # Kreis-Header
                st.markdown(f"""
                <div style="border-left:4px solid {farbe};padding:0.4rem 0.8rem;
                            background:#F8F9FA;border-radius:0 6px 6px 0;margin-bottom:0.6rem;">
                    <strong style="color:{farbe};">Kreis {ki["kreis_nr"]}</strong> ·
                    Verdampfung <strong>{ki["verdampfung_c"]}°C</strong> ·
                    Gesamtleistung <strong>{ki["leistung_kw_gesamt"]} kW</strong> ·
                    {ki["anzahl_direkt"]} Direkt / {ki["anzahl_extern"]} Extern
                </div>
                """, unsafe_allow_html=True)

                # Schaltschrank + Maschine zuweisen
                sk_c1, sk_c2, sk_c3 = st.columns(3)
                with sk_c1:
                    st.text_input("Schaltschrank / Verteiler",
                        placeholder="z.B. SK-01, Verteiler Nord",
                        key=f"sk_{ki_idx}_{ki['kreis_nr']}",
                        help="Zentraler Schaltschrank oder Verteiler für diesen Kreis")
                with sk_c2:
                    st.text_input("Maschine / Verbundanlage",
                        placeholder="z.B. VA-1, Aggregat TK",
                        key=f"ma_{ki_idx}_{ki['kreis_nr']}",
                        help="Kältemaschine oder Verbundanlage die diesen Kreis versorgt")
                with sk_c3:
                    ms_opts2 = st.session_state.get("maschinenstandorte",[])
                    if ms_opts2:
                        ms_labels2 = [m.get("name","") for m in ms_opts2]
                        st.selectbox("Maschinenstandort",
                            ms_labels2, key=f"ms_{ki_idx}_{ki['kreis_nr']}",
                            help="Wo steht die Maschine für diesen Kreis?")

                # Kühlstellen-Tabelle pro Kreis
                import pandas as pd
                rows_kreis = []
                for ks2 in ki["kuehlstellen_alle"]:
                    rows_kreis.append({
                        "Kühlstelle": ks2["name"],
                        "Lieferumfang": "🔧 Direkt" if ks2.get("lieferumfang")=="direkt" else "🔄 Extern",
                        "Raumtemp. °C": ks2.get("raum_temp_soll_c","–"),
                        "Leistung kW": ks2.get("kaelteleistung_kw","–"),
                        "Länge [m]": ks2.get("leitungslaenge_m","–"),
                        "Standort": ks2.get("standort_raum","–"),
                        "Regler": ks2.get("komponenten",{}).get("schaltkasten",{}).get(
                            "parameter",{}).get("regler_typ","–")[:25] if ks2.get("komponenten",{}).get("schaltkasten") else "–",
                    })
                st.dataframe(pd.DataFrame(rows_kreis), use_container_width=True,
                             hide_index=True, height=min(150+len(rows_kreis)*35, 400))

                # Direktleistung vs Extern
                if ki["anzahl_direkt"] > 0:
                    st.caption(f"🔧 Direkt: {', '.join(ki['kuehlstellen_direkt'])}")
                if ki["anzahl_extern"] > 0:
                    st.caption(f"🔄 Extern: {', '.join(ki['kuehlstellen_extern'])}")

# ============================================================
# TAB 4 – KABELPLANUNG
# ============================================================
with t4:
    st.markdown('<div class="sec">🔌 Strukturierte Kabelliste – pro Verbraucher</div>', unsafe_allow_html=True)
    st.caption("Geordnet nach Kreis → Kühlstelle → Kabeltyp. Von/Bis und Länge editierbar. Excel-Export für Partner-Review und Re-Import.")

    ks_list3 = st.session_state.kuehlstellen
    if not ks_list3:
        st.info("Zuerst Kühlstellen in Tab 2 erfassen.")
    else:
        # HACCP-Module konfigurieren
        with st.expander("📋 HACCP-Module konfigurieren", expanded=False):
            st.caption("HACCP-Module können an einer Kühlstelle oder zentral sitzen und 1–6 Fühler aufnehmen. Von dort Bus zum Monitoring-System.")
            hm_list = st.session_state.get("haccp_module", [])
            MONITORING_ZIELE = ["Xweb / Dixell","Kiconex Gateway","Frigodata Gateway","Carel boss","Regler-intern","Sonstiges"]
            for hm_i, hm in enumerate(hm_list):
                hmc1,hmc2,hmc3,hmc4,hmc5 = st.columns([2,1,2,2,1])
                with hmc1:
                    hm["bezeichnung"] = st.text_input("Bezeichnung", value=hm.get("bezeichnung",""),
                        key=f"hm_bez_{hm_i}", placeholder="z.B. HACCP-Modul EG")
                with hmc2:
                    hm["max_fuehler"] = st.number_input("Max. Fühler", 1, 6, int(hm.get("max_fuehler",4)),
                        key=f"hm_f_{hm_i}")
                with hmc3:
                    hm["standort"] = st.text_input("Standort", value=hm.get("standort",""),
                        key=f"hm_st_{hm_i}", placeholder="z.B. an Kühlstelle KR-01, Maschinenraum")
                with hmc4:
                    mz_idx = MONITORING_ZIELE.index(hm.get("monitoring_ziel","Xweb / Dixell")) if hm.get("monitoring_ziel") in MONITORING_ZIELE else 0
                    hm["monitoring_ziel"] = st.selectbox("Bus → Monitoring", MONITORING_ZIELE,
                        index=mz_idx, key=f"hm_mz_{hm_i}")
                    hm["laenge_zum_monitoring"] = st.number_input("Länge [m]", 0, 500,
                        int(hm.get("laenge_zum_monitoring",20)), key=f"hm_lm_{hm_i}")
                with hmc5:
                    if st.button("🗑️", key=f"hm_del_{hm_i}"):
                        hm_list.pop(hm_i); st.session_state.haccp_module = hm_list; st.rerun()
                hm_list[hm_i] = hm
            if st.button("➕ HACCP-Modul hinzufügen"):
                hm_list.append({"bezeichnung":"HACCP-Modul","max_fuehler":4,"standort":"","monitoring_ziel":"Xweb / Dixell","laenge_zum_monitoring":20})
                st.rerun()
            st.session_state.haccp_module = hm_list

        st.markdown("---")

        # Kabelliste generieren
        if st.button("🔄 Kabelliste neu berechnen", type="primary"):
            st.session_state.kabelliste_cache = None

        if st.session_state.get("kabelliste_cache") is None:
            with st.spinner("Berechne Kabelliste..."):
                kl = erzeuge_kabelliste(
                    ks_list3,
                    st.session_state.get("maschinenstandorte",[]),
                    st.session_state.get("steuerung",""),
                    st.session_state.get("haccp_module",[])
                )
                st.session_state.kabelliste_cache = kl

        kl = st.session_state.kabelliste_cache or []

        if kl:
            df_kl = pd.DataFrame(kl)

            # Statistik
            sc1,sc2,sc3,sc4 = st.columns(4)
            sc1.metric("Kabel gesamt", len(kl))
            sc2.metric("Direkt", len([k for k in kl if k["Lieferumfang"]=="Direkt"]))
            sc3.metric("Extern", len([k for k in kl if k["Lieferumfang"]=="Extern"]))
            total_m = round(sum(float(k.get("Länge [m]",0) or 0) for k in kl),1)
            sc4.metric("Meter gesamt", f"{total_m} m")

            # Gefilterte Ansicht
            col_filter1, col_filter2 = st.columns(2)
            with col_filter1:
                filter_kreis = st.multiselect("Kreis filtern",
                    options=sorted(df_kl["Kreis Nr."].unique()),
                    format_func=lambda x: f"Kreis {x}",
                    default=[])
            with col_filter2:
                filter_lf = st.multiselect("Lieferumfang",
                    ["Direkt","Extern"], default=[])

            df_anzeige = df_kl.copy()
            if filter_kreis:
                df_anzeige = df_anzeige[df_anzeige["Kreis Nr."].isin(filter_kreis)]
            if filter_lf:
                df_anzeige = df_anzeige[df_anzeige["Lieferumfang"].isin(filter_lf)]

            # Tabelle anzeigen (ohne _id)
            anzeige_cols = ["Kreis Nr.","Kühlstelle / Ort","Lieferumfang","Kategorie",
                           "Bezeichnung / Verbraucher","Kabeltyp","Querschnitt [mm²]",
                           "Adern","Von","Bis","Länge [m]","Norm","Bemerkung"]
            st.dataframe(df_anzeige[anzeige_cols], use_container_width=True, height=500,
                column_config={
                    "Kreis Nr.": st.column_config.NumberColumn(width=70),
                    "Lieferumfang": st.column_config.TextColumn(width=80),
                    "Kategorie": st.column_config.TextColumn(width=90),
                    "Länge [m]": st.column_config.NumberColumn(format="%.1f m", width=80),
                    "Von": st.column_config.TextColumn(width=180),
                    "Bis": st.column_config.TextColumn(width=180),
                })

            # ZUSAMMENFASSUNG
            st.markdown('<div class="sec">📊 Kabelzusammenfassung – Gesamtmeter je Kabeltyp</div>', unsafe_allow_html=True)
            summen = kabelliste_zusammenfassung(kl)
            df_sum = pd.DataFrame(summen)
            st.dataframe(df_sum, use_container_width=True, hide_index=True,
                column_config={
                    "Gesamt [m]": st.column_config.NumberColumn(format="%.1f m"),
                    "Direkt [m]": st.column_config.NumberColumn(format="%.1f m"),
                    "Extern [m]": st.column_config.NumberColumn(format="%.1f m"),
                })

            # EXPORT
            st.markdown("---")
            st.markdown('<div class="sec">📥 Export & Partner-Review</div>', unsafe_allow_html=True)
            st.caption("Excel exportieren → Partner prüft + korrigiert → Excel wieder importieren → Merge mit Projekt")

            p_nr = st.session_state.projekt.get("nummer","projekt")
            ex1,ex2,ex3 = st.columns(3)

            with ex1:
                # CSV Export
                csv_kl = df_kl.to_csv(index=False, sep=";", encoding="utf-8-sig")
                st.download_button("📥 CSV Kabelliste",
                    data=csv_kl.encode("utf-8-sig"),
                    file_name=f"coolWIRE_{p_nr}_Kabelliste.csv",
                    mime="text/csv", use_container_width=True,
                    help="CSV für einfache Weiterverarbeitung")

            with ex2:
                # Excel Export via openpyxl
                import io
                try:
                    from openpyxl import Workbook
                    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                    from openpyxl.utils import get_column_letter

                    wb = Workbook()

                    # Sheet 1: Kabelliste
                    ws = wb.active
                    ws.title = "Kabelliste"

                    # Header
                    headers = ["ID","Kreis Nr.","Kreis","Kühlstelle / Ort","Lieferumfang",
                               "Kategorie","Bezeichnung / Verbraucher","Kabeltyp",
                               "Querschnitt [mm²]","Adern","Von","Bis","Länge [m]","Norm","Bemerkung"]
                    FARBEN = {1:"4472C4",2:"ED7D31",3:"A9D18E",4:"9B59B6",5:"E74C3C"}
                    hdr_fill = PatternFill("solid", start_color="1F3864")
                    hdr_font = Font(bold=True, color="FFFFFF", name="Arial", size=9)
                    for col, h in enumerate(headers, 1):
                        cell = ws.cell(row=1, column=col, value=h)
                        cell.fill = hdr_fill
                        cell.font = hdr_font
                        cell.alignment = Alignment(horizontal="center", wrap_text=True)

                    # Daten
                    prev_kreis = None
                    for row_i, zeile in enumerate(kl, 2):
                        kreis_nr = zeile.get("Kreis Nr.",1)
                        farbe = FARBEN.get(kreis_nr, "D9D9D9")
                        row_fill = PatternFill("solid", start_color=f"E8F0F{kreis_nr % 5 + 1}" if kreis_nr < 5 else "F2F2F2")

                        vals = [zeile.get("_id",""), kreis_nr, zeile.get("Kreis",""),
                                zeile.get("Kühlstelle / Ort",""), zeile.get("Lieferumfang",""),
                                zeile.get("Kategorie",""), zeile.get("Bezeichnung / Verbraucher",""),
                                zeile.get("Kabeltyp",""), zeile.get("Querschnitt [mm²]",""),
                                zeile.get("Adern",""), zeile.get("Von",""), zeile.get("Bis",""),
                                zeile.get("Länge [m]",0), zeile.get("Norm",""), zeile.get("Bemerkung","")]

                        for col, val in enumerate(vals, 1):
                            cell = ws.cell(row=row_i, column=col, value=val)
                            cell.font = Font(name="Arial", size=9)
                            if col in [11,12,13,15]:  # Von, Bis, Länge, Bemerkung = editierbar
                                cell.fill = PatternFill("solid", start_color="FFFDE7")
                            if prev_kreis and prev_kreis != kreis_nr:
                                thin = Side(style="thin", color="4472C4")
                                cell.border = Border(top=thin)
                        prev_kreis = kreis_nr

                    # Spaltenbreiten
                    widths = [8,7,25,30,10,12,40,15,12,7,25,25,10,20,30]
                    for i, w in enumerate(widths, 1):
                        ws.column_dimensions[get_column_letter(i)].width = w
                    ws.freeze_panes = "A2"
                    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

                    # Sheet 2: Zusammenfassung
                    ws2 = wb.create_sheet("Zusammenfassung")
                    sum_headers = ["Kabeltyp","Querschnitt [mm²]","Adern","Kategorie",
                                   "Anzahl Leitungen","Gesamt [m]","Direkt [m]","Extern [m]"]
                    for col, h in enumerate(sum_headers, 1):
                        cell = ws2.cell(row=1, column=col, value=h)
                        cell.fill = hdr_fill
                        cell.font = hdr_font
                    for row_i, s in enumerate(summen, 2):
                        vals2 = [s["Kabeltyp"],s["Querschnitt [mm²]"],s["Adern"],
                                 s["Kategorie"],s["Anzahl Leitungen"],
                                 s["Gesamt [m]"],s["Direkt [m]"],s["Extern [m]"]]
                        for col, val in enumerate(vals2, 1):
                            ws2.cell(row=row_i, column=col, value=val).font = Font(name="Arial",size=9)
                    ws2.column_dimensions["A"].width = 20
                    ws2.column_dimensions["D"].width = 15

                    # Sheet 3: Projekt-Info
                    ws3 = wb.create_sheet("Projektinfo")
                    p3 = st.session_state.projekt
                    infos = [
                        ("Projekt", p3.get("name","")),
                        ("Nr.", p3.get("nummer","")),
                        ("Kunde", p3.get("kunde","")),
                        ("Standort", p3.get("standort","")),
                        ("Bearbeiter", p3.get("bearbeiter","")),
                        ("Datum", p3.get("erstellt","")),
                        ("Generiert", datetime.now().strftime("%d.%m.%Y %H:%M")),
                        ("Kühlstellen", len(ks_list3)),
                        ("Kabel gesamt", len(kl)),
                        ("Meter gesamt", total_m),
                    ]
                    for r, (k,v) in enumerate(infos, 1):
                        ws3.cell(row=r,column=1,value=k).font = Font(bold=True,name="Arial",size=9)
                        ws3.cell(row=r,column=2,value=v).font = Font(name="Arial",size=9)

                    buf = io.BytesIO()
                    wb.save(buf)
                    buf.seek(0)
                    st.download_button("📥 Excel Kabelliste",
                        data=buf.getvalue(),
                        file_name=f"coolWIRE_{p_nr}_Kabelliste.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        help="Gelbe Felder = editierbar (Von, Bis, Länge, Bemerkung)")
                except Exception as ex:
                    st.error(f"Excel-Export Fehler: {ex}")

            with ex3:
                # Excel Re-Import
                xl_reimport = st.file_uploader("📂 Excel Re-Import (nach Partner-Review)",
                    type=["xlsx","xls"], key="kl_reimport", label_visibility="collapsed")
                if xl_reimport:
                    try:
                        df_imp = pd.read_excel(xl_reimport, sheet_name="Kabelliste")
                        # Längen + Von/Bis + Bemerkung zurückschreiben
                        updated = 0
                        for _, row in df_imp.iterrows():
                            id_col = str(row.get("ID",""))
                            for zeile in kl:
                                if zeile.get("_id") == id_col:
                                    zeile["Von"] = str(row.get("Von", zeile["Von"]))
                                    zeile["Bis"] = str(row.get("Bis", zeile["Bis"]))
                                    zeile["Länge [m]"] = float(row.get("Länge [m]", zeile["Länge [m]"]) or 0)
                                    zeile["Bemerkung"] = str(row.get("Bemerkung","") or "")
                                    updated += 1
                        st.session_state.kabelliste_cache = kl
                        st.success(f"✅ {updated} Zeilen aktualisiert!")
                        st.rerun()
                    except Exception as ex:
                        st.error(f"Import-Fehler: {ex}")

        else:
            st.info("Keine Kabelliste – Kühlstellen konfigurieren und 'Kabelliste neu berechnen' klicken.")

    st.markdown("---")
    st.markdown('<div class="sec">⚡ Querschnitt berechnen</div>', unsafe_allow_html=True)
    qc1,qc2 = st.columns(2, gap="large")
    with qc1:
        qph = st.radio("Phasenzahl",[1,3], format_func=lambda x:"1-phasig (230V)" if x==1 else "3-phasig (400V)", horizontal=True)
        qlw = st.number_input("Leistung [W]",100,500000,3000,100)
        qla = st.number_input("Länge [m]",1.0,2000.0,25.0,1.0)
        qcp = st.slider("cos φ",0.5,1.0,0.9,0.01)
        if st.button("🔢 Berechnen", type="primary", use_container_width=True):
            st.session_state.qs_result = berechne_leitungsquerschnitt(qlw,230 if qph==1 else 400,qla,qcp,phasen=qph)
    with qc2:
        if st.session_state.qs_result:
            r = st.session_state.qs_result
            st.metric("Gewählter Querschnitt", f"{r['querschnitt_gewaehlt_mm2']} mm²")
            st.metric("Betriebsstrom", f"{r['strom_a']} A")
            st.caption(f"Berechnet: {r['querschnitt_berechnet_mm2']} mm²")
            st.caption(f"⚠️ {r['hinweis']}")

# ============================================================
# TAB 5 – DOKUMENTATION
# ============================================================
# ============================================================
# TAB 5 – ROHRNETZ
# ============================================================
with t5_rohr:
    st.markdown('<div class="sec">🔧 Kältetechnische Rohrnetzberechnung</div>', unsafe_allow_html=True)
    st.info("🚧 **Rohrnetzberechnung – in Entwicklung**\n\nDieser Tab wird mit coolROHR v6.2 erweitert. "
            "Die Kühlstellendaten, Kälteleistungen, Leitungslängen und Kältemittel werden automatisch aus dem Projekt übernommen.")

    ks_rohr = st.session_state.kuehlstellen
    if not ks_rohr:
        st.warning("Zuerst Kühlstellen in Tab 2 erfassen.")
    else:
        from collections import defaultdict
        kreise_rohr = defaultdict(list)
        for ks in ks_rohr:
            kreise_rohr[ks.get("kreis",0)].append(ks)

        kreis_farben = {1:"#DDEBF7", 2:"#E2EFDA", 3:"#FFF2CC"}
        kreis_au = {
            1: {"modell":"MDV-SY-50582 #2","kaeltemittel":"R513A","verdampfung_c":-8,"kondensation_c":47},
            2: {"modell":"Sigilus BDF-NG-1075","kaeltemittel":"R449A","verdampfung_c":-28,"kondensation_c":42},
            3: {"modell":"MDV-SY-50582 #1","kaeltemittel":"R513A","verdampfung_c":-8,"kondensation_c":47},
        }

        st.markdown("### Kreisübersicht – Eingangsdaten für Rohrnetzberechnung")
        for kr_nr in sorted(kreise_rohr.keys()):
            ks_liste = kreise_rohr[kr_nr]
            au = kreis_au.get(kr_nr, {})
            kw_sum = round(sum(k.get("kaelteleistung_kw",0) for k in ks_liste), 2)
            laengen = [k.get("leitungslaenge_m",0) for k in ks_liste if k.get("leitungslaenge_m")]

            with st.expander(f"Kreis {kr_nr} – {au.get('modell','?')} · {au.get('kaeltemittel','?')} · {kw_sum} kW · {len(ks_liste)} Kühlstellen", expanded=True):
                rc1, rc2, rc3, rc4 = st.columns(4)
                rc1.metric("Kältemittel", au.get("kaeltemittel","?"))
                rc2.metric("Verdampfung", f"{au.get('verdampfung_c','?')}°C")
                rc3.metric("Kondensation", f"{au.get('kondensation_c','?')}°C")
                rc4.metric("Gesamtleistung", f"{kw_sum} kW")

                st.markdown("**Kühlstellen in diesem Kreis:**")
                rohr_data = []
                for ks in ks_liste:
                    rohr_data.append({
                        "E-Nr": ks.get("e_nr","–"),
                        "Pos.": ks.get("pos_nr","–"),
                        "Bezeichnung": ks.get("name",""),
                        "Leistung [kW]": ks.get("kaelteleistung_kw","?"),
                        "Länge AU→KS [m]": ks.get("leitungslaenge_m","?"),
                        "Temp [°C]": ks.get("raum_temp_soll_c","?"),
                    })
                import pandas as pd
                st.dataframe(pd.DataFrame(rohr_data), use_container_width=True, hide_index=True)

                st.markdown("---")
                st.markdown("**⚙️ Rohrnetzberechnung** *(coolROHR Integration – folgt)*")
                rr1, rr2, rr3 = st.columns(3)
                with rr1:
                    st.markdown("**Saugleitung**")
                    st.caption("Rohrdimension: wird berechnet")
                    st.caption("Druckabfall: wird berechnet")
                with rr2:
                    st.markdown("**Druckleitung**")
                    st.caption("Rohrdimension: wird berechnet")
                    st.caption("Druckabfall: wird berechnet")
                with rr3:
                    st.markdown("**Flüssigkeitsleitung**")
                    st.caption("Rohrdimension: wird berechnet")
                    st.caption("Druckabfall: wird berechnet")

                st.info(f"📋 Sobald coolROHR integriert ist werden hier die Rohrdimensionen für alle "
                        f"{len(ks_liste)} Kühlstellen in Kreis {kr_nr} automatisch berechnet.")

with t5:
    st.markdown('<div class="sec">📄 Projektdokumentation & Export</div>', unsafe_allow_html=True)
    p2 = st.session_state.projekt
    ks4 = st.session_state.kuehlstellen
    dc1,dc2,dc3,dc4 = st.columns(4)
    dc1.metric("Projekt", p2.get("name","–"))
    dc2.metric("Kunde", p2.get("kunde","–"))
    dc3.metric("Kühlstellen", len(ks4))
    dc4.metric("Steuerung", st.session_state.steuerung or "–")

    if ks4:
        st.markdown("---")
        st.markdown('<div class="sec">📊 Kühlstellenübersicht</div>', unsafe_allow_html=True)
        df_d = pd.DataFrame([{
            "Nr.": k["nummer"], "Name": k["name"],
            "Bereich": TEMP_BEREICHE.get(k.get("temp_bereich","NK"),{}).get("label","–"),
            "Raumtemp.°C": k.get("raum_temp_soll_c","–"),
            "Verdampfung°C": k.get("verdampfung_custom_c","–"),
            "kW": k.get("kaelteleistung_kw","–"),
            "Kreis": k.get("kreis",1),
            "Lieferumfang": "Direkt" if k.get("lieferumfang")=="direkt" else "Extern",
            "Komponenten": len([v for v in k.get("komponenten",{}).values() if v.get("aktiv")]),
            "Notizen": k.get("notizen","")
        } for k in ks4])
        st.dataframe(df_d, use_container_width=True, hide_index=True)
        st.markdown("---")
        ex1,ex2,ex3 = st.columns(3)
        with ex1:
            st.download_button("📥 Kühlstellen CSV",
                data=df_d.to_csv(index=False,sep=";",encoding="utf-8-sig").encode("utf-8-sig"),
                file_name=f"coolWIRE_{p2.get('nummer','')}_Kuehlstellen.csv",
                mime="text/csv", use_container_width=True)
        with ex2:
            exp2 = {"version":"2.0","generiert":datetime.now().isoformat(),
                    "projekt":p2,"kuehlstellen":ks4,"steuerung":st.session_state.steuerung}
            st.download_button("📥 Projekt JSON",
                data=json.dumps(exp2,ensure_ascii=False,indent=2).encode("utf-8"),
                file_name=f"coolWIRE_{p2.get('nummer','')}_komplett.json",
                mime="application/json", use_container_width=True)
        with ex3:
            df_mx = exportiere_gesamtliste()
            st.download_button("📥 Kabelmatrix CSV",
                data=df_mx.to_csv(index=False,sep=";",encoding="utf-8-sig").encode("utf-8-sig"),
                file_name="coolWIRE_Kabelmatrix.csv",
                mime="text/csv", use_container_width=True)

# ============================================================
# ============================================================
# TAB 6 – ADMIN
# ============================================================
if ist_admin() and t6:
    with t6:
        admin_tab1, admin_tab2, admin_tab3 = st.tabs([
            "👥 Nutzerverwaltung",
            "🎛️ Regler-Datenbank",
            "🎨 Theme & Design"
        ])

        # ── NUTZERVERWALTUNG ──
        with admin_tab1:
            zeige_nutzerverwaltung()

        # ── REGLER-DATENBANK ──
        with admin_tab2:
            st.markdown('<div class="sec">🎛️ Regler-Datenbank</div>', unsafe_allow_html=True)
            st.caption("Hier kannst du neue Reglertypen hinzufügen. Die App lernt sie und verwendet sie in allen Kühlstellen-Dropdowns.")

            db = st.session_state.db
            regler_liste = get_regler_liste(db)

            # Bestehende Regler anzeigen
            if regler_liste:
                import pandas as pd
                df_reg = pd.DataFrame([{
                    "ID": r.get("id"),
                    "Hersteller": r.get("hersteller"),
                    "Modell": r.get("modell"),
                    "Typ": r.get("typ"),
                    "Spannung": r.get("spannung"),
                    "Bus": r.get("bus"),
                    "Relais": r.get("relais_ausgaenge"),
                    "Fühler": r.get("fuehler_eingaenge"),
                    "Einsatz": ", ".join(r.get("einsatzgebiet",[])),
                    "Angelegt": r.get("angelegt_am",""),
                    "Von": r.get("angelegt_von","")
                } for r in regler_liste])
                st.dataframe(df_reg, use_container_width=True, hide_index=True, height=280)

                # Regler löschen (nur wenn > 10 vorhanden = eigene Einträge)
                if len(regler_liste) > 10:
                    custom_regler = regler_liste[10:]
                    del_opt = st.selectbox("Eigenen Regler löschen",
                        options=[r["id"] for r in custom_regler],
                        format_func=lambda x: next((f"{r['hersteller']} {r['modell']}" for r in custom_regler if r["id"]==x), x))
                    if st.button("🗑️ Regler löschen", type="secondary"):
                        db = delete_regler(db, del_opt)
                        speichere_datenbank(db)
                        st.session_state.db = db
                        st.success("Regler gelöscht!")
                        st.rerun()

            st.markdown("---")
            st.markdown("**➕ Neuen Regler hinzufügen**")
            st.caption("Der neue Regler wird sofort in allen Kühlstellen-Dropdowns verfügbar.")

            with st.form("neuer_regler_form", clear_on_submit=True):
                nr1, nr2 = st.columns(2)
                with nr1:
                    n_hersteller = st.text_input("Hersteller *", placeholder="z.B. Danfoss, Carel, Wurm...")
                    n_modell     = st.text_input("Modell *", placeholder="z.B. EKC 202, IR33, UDO-4")
                    n_typ        = st.selectbox("Typ", ["Kühlstellenregler", "Verbundregler / Rack Controller",
                                                         "Supervisor / Monitoring", "EEV-Regler", "Sonstiger"])
                    n_spannung   = st.selectbox("Spannung", ["230V 1-phasig", "24V AC/DC", "400V 3-phasig"])
                with nr2:
                    n_bus        = st.text_input("Bus / Protokoll", placeholder="z.B. RS485 Modbus, CAN-Bus, pLAN")
                    n_display    = st.text_input("Display", placeholder="z.B. LED 4-stellig, LCD Grafik")
                    n_relais     = st.number_input("Relaisausgänge", 0, 20, 3)
                    n_fuehler    = st.number_input("Fühlereingänge", 0, 20, 2)

                n_einsatz   = st.multiselect("Einsatzgebiet", ["NK", "TK", "NK+", "TK+", "HNK", "Verbund", "Industrie"],
                                             default=["NK"])
                n_besond    = st.text_area("Besonderheiten / Notizen",
                                           placeholder="z.B. EEV-Ansteuerung, HACCP-fähig, besondere Eigenschaft...",
                                           height=70)
                submitted = st.form_submit_button("✅ Regler speichern", type="primary", use_container_width=True)

            if submitted:
                if not n_hersteller or not n_modell:
                    st.error("Hersteller und Modell sind Pflichtfelder.")
                else:
                    db = add_regler(db, n_hersteller, n_modell, n_typ, n_spannung, n_bus,
                                   n_display, int(n_relais), int(n_fuehler), n_besond,
                                   n_einsatz, get_display_name())
                    speichere_datenbank(db)
                    st.session_state.db = db
                    st.success(f"✅ {n_hersteller} {n_modell} gespeichert und sofort verfügbar!")
                    st.rerun()

            # DB Export/Import
            st.markdown("---")
            st.markdown("**📦 Datenbank sichern / wiederherstellen**")
            col_exp1, col_exp2 = st.columns(2)
            with col_exp1:
                import json as json_mod
                st.download_button("📥 Datenbank exportieren (JSON)",
                    data=json_mod.dumps(db, ensure_ascii=False, indent=2).encode("utf-8"),
                    file_name="coolwire_datenbank.json",
                    mime="application/json", use_container_width=True)
            with col_exp2:
                db_upload = st.file_uploader("📂 Datenbank importieren", type=["json"],
                                             key="db_import", label_visibility="collapsed")
                if db_upload:
                    try:
                        imported = json_mod.load(db_upload)
                        speichere_datenbank(imported)
                        st.session_state.db = imported
                        st.success("Datenbank importiert!")
                        st.rerun()
                    except Exception as e:
                        st.error(f"Fehler: {e}")

        # ── THEME & DESIGN ──
        with admin_tab3:
            st.markdown('<div class="sec">🎨 Theme & Design</div>', unsafe_allow_html=True)
            st.caption("Farben, Schrift und Design der App anpassen. Änderungen werden sofort gespeichert und beim nächsten Start übernommen.")

            db = st.session_state.db
            current_theme = get_theme(db)
            presets = get_theme_presets(db)

            # Preset auswählen
            st.markdown("**🎭 Preset auswählen**")
            preset_namen = {k: v["name"] for k,v in presets.items()}
            preset_keys = list(preset_namen.keys())
            selected_preset = st.selectbox("Design-Preset", preset_keys,
                format_func=lambda x: preset_namen[x])

            if st.button("✅ Preset anwenden", type="primary"):
                preset_data = presets[selected_preset].copy()
                new_theme = current_theme.copy()
                new_theme.update(preset_data)
                new_theme["name"] = preset_data["name"]
                db = speichere_theme(db, new_theme)
                speichere_datenbank(db)
                st.session_state.db = db
                ok = schreibe_config_toml(new_theme)
                if ok:
                    st.success(f"✅ Theme '{preset_data['name']}' gespeichert in config.toml – **BAT neu starten** damit es wirksam wird!")
                else:
                    st.warning("Theme gespeichert aber config.toml konnte nicht geschrieben werden – manuell anpassen.")
                st.rerun()

            st.markdown("---")
            st.markdown("**🎨 Farben individuell anpassen**")

            tc1, tc2, tc3 = st.columns(3)
            with tc1:
                new_primary = st.color_picker("Hauptfarbe (Buttons, Tabs, Akzente)",
                    value=current_theme.get("primary_color","#36A9E1"))
                new_dark = st.color_picker("Textfarbe / Dark",
                    value=current_theme.get("dark_color","#3C3C3B"))
                new_accent = st.color_picker("Akzentfarbe (Links, Hover)",
                    value=current_theme.get("accent_color","#0078B8"))
            with tc2:
                new_light = st.color_picker("Hintergrundfarbe (Hell)",
                    value=current_theme.get("light_bg","#F4F8FC"))
                new_border = st.color_picker("Rahmen / Border",
                    value=current_theme.get("border_color","#D0E8F5"))
                new_grad_start = st.color_picker("Header Gradient Start",
                    value=current_theme.get("header_gradient_start","#1a6fa8"))
                new_grad_end = st.color_picker("Header Gradient Ende",
                    value=current_theme.get("header_gradient_end","#5bc4f5"))
            with tc3:
                font_opts = ["DM Sans", "Inter", "Roboto", "Open Sans", "Source Sans Pro",
                             "Nunito", "Montserrat", "Lato", "Poppins"]
                new_font = st.selectbox("Schriftart",
                    font_opts,
                    index=font_opts.index(current_theme.get("font_family","DM Sans"))
                          if current_theme.get("font_family","DM Sans") in font_opts else 0)
                new_fontsize = st.selectbox("Schriftgröße",
                    ["12px","13px","14px","15px","16px"],
                    index=["12px","13px","14px","15px","16px"].index(
                        current_theme.get("font_size_base","14px"))
                          if current_theme.get("font_size_base","14px") in ["12px","13px","14px","15px","16px"] else 2)
                new_radius = st.selectbox("Eckenrundung",
                    ["0px","4px","6px","8px","10px","14px","20px"],
                    index=["0px","4px","6px","8px","10px","14px","20px"].index(
                        current_theme.get("border_radius","10px"))
                          if current_theme.get("border_radius","10px") in ["0px","4px","6px","8px","10px","14px","20px"] else 4)

            # Vorschau
            st.markdown("---")
            st.markdown("**👁️ Vorschau**")
            st.markdown(f"""
            <div style="background:linear-gradient(135deg,{new_grad_start},{new_grad_end});
                        padding:1rem 1.5rem;border-radius:{new_radius};margin-bottom:0.8rem;">
                <span style="color:white;font-family:{new_font},sans-serif;font-size:1.3rem;font-weight:700;">
                    ⚡ °coolWIRE – Vorschau Header
                </span>
            </div>
            <div style="background:{new_light};border:1.5px solid {new_border};
                        border-radius:{new_radius};padding:1rem;margin-bottom:0.5rem;">
                <span style="color:{new_dark};font-family:{new_font},sans-serif;">
                    Beispiel-Karte mit Text · Hauptfarbe:
                    <span style="color:{new_primary};font-weight:700;">●</span> {new_primary}
                </span>
            </div>
            """, unsafe_allow_html=True)

            if st.button("💾 Theme speichern & config.toml aktualisieren", type="primary", use_container_width=True):
                new_theme = {
                    "name": "Individuell",
                    "primary_color": new_primary,
                    "dark_color": new_dark,
                    "light_bg": new_light,
                    "border_color": new_border,
                    "accent_color": new_accent,
                    "success_color": current_theme.get("success_color","#27AE60"),
                    "warning_color": current_theme.get("warning_color","#E67E22"),
                    "error_color": current_theme.get("error_color","#E74C3C"),
                    "font_family": new_font,
                    "font_size_base": new_fontsize,
                    "border_radius": new_radius,
                    "card_shadow": current_theme.get("card_shadow","0 2px 8px rgba(0,0,0,0.06)"),
                    "header_gradient_start": new_grad_start,
                    "header_gradient_end": new_grad_end,
                }
                db = speichere_theme(db, new_theme)
                speichere_datenbank(db)
                st.session_state.db = db
                ok = schreibe_config_toml(new_theme)
                if ok:
                    st.success("✅ Gespeichert! Jetzt **BAT neu starten** – dann ist das neue Theme aktiv.")
                else:
                    st.warning("Gespeichert, aber config.toml konnte nicht geschrieben werden.")
                st.rerun()

            st.markdown("---")
            st.markdown("**🌗 Schnellauswahl Hell / Dunkel:**")
            col_hd1, col_hd2 = st.columns(2)
            with col_hd1:
                if st.button("☀️ Hell (Light Mode)", use_container_width=True):
                    hell_theme = DEFAULT_DB["theme_presets"]["coolsulting_standard"].copy()
                    hell_theme["name"] = "Hell"
                    db = speichere_theme(db, hell_theme)
                    speichere_datenbank(db)
                    st.session_state.db = db
                    schreibe_config_toml(hell_theme)
                    st.success("✅ Light Mode – BAT neu starten!")
                    st.rerun()
            with col_hd2:
                if st.button("🌙 Dunkel (Dark Mode)", use_container_width=True):
                    dunkel_theme = DEFAULT_DB["theme_presets"]["dunkel"].copy()
                    dunkel_theme["name"] = "Dunkel"
                    db = speichere_theme(db, dunkel_theme)
                    speichere_datenbank(db)
                    st.session_state.db = db
                    schreibe_config_toml(dunkel_theme)
                    st.success("✅ Dark Mode – BAT neu starten!")
                    st.rerun()

            if st.button("🔄 Auf Standard zurücksetzen"):
                std = DEFAULT_DB["theme"].copy()
                db = speichere_theme(db, std)
                speichere_datenbank(db)
                st.session_state.db = db
                schreibe_config_toml(std)
                st.success("Zurückgesetzt – BAT neu starten!")
                st.rerun()

st.markdown('<div class="footer">⚡ °coolWIRE v2.0 · coolsulting e.U. · Michael Schäpers · Mozartstraße 11, 4020 Linz · ATU78952901</div>', unsafe_allow_html=True)
