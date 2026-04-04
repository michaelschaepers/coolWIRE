# modules/ki_import.py
# °coolWIRE v2.1 – KI-Import: PDF / DWG / Excel / Screenshot / Text + Lernfunktion
# (c) coolsulting e.U. | Michael Schäpers

import streamlit as st
import base64
import json
import re
import io

try:
    import anthropic
    ANTHROPIC_OK = True
except ImportError:
    ANTHROPIC_OK = False

try:
    import ezdxf
    EZDXF_OK = True
except ImportError:
    EZDXF_OK = False

try:
    import openpyxl
    EXCEL_OK = True
except ImportError:
    EXCEL_OK = False

def get_api_key():
    try:
        return st.secrets["anthropic"]["api_key"]
    except Exception:
        return None

def api_verfuegbar():
    return ANTHROPIC_OK and bool(get_api_key())

def _client():
    return anthropic.Anthropic(api_key=get_api_key())

# --- LERNFUNKTION ---

LERN_KEY = "ki_gelerntes"

def lerne_aus_ergebnis(ki_result):
    if "fehler" in ki_result:
        return 0
    if LERN_KEY not in st.session_state:
        st.session_state[LERN_KEY] = []
    neu = []
    for ks in ki_result.get("kuehlstellen", []):
        muster = {
            "name_pattern": ks.get("name","").lower()[:30],
            "temp_bereich": ks.get("temp_bereich","NK"),
            "raum_temp_soll_c": ks.get("raum_temp_soll_c"),
            "verdampfung_c": ks.get("verdampfung_c"),
            "komponenten_erkannt": ks.get("komponenten_erkannt",[]),
        }
        bereits = [m["name_pattern"] for m in st.session_state[LERN_KEY]]
        if muster["name_pattern"] not in bereits:
            neu.append(muster)
    st.session_state[LERN_KEY].extend(neu)
    return len(neu)

def get_gelernte_muster():
    return st.session_state.get(LERN_KEY, [])

def lernmuster_als_kontext():
    muster = get_gelernte_muster()
    if not muster:
        return ""
    lines = ["GELERNTE MUSTER AUS FRUEHEREN PROJEKTEN:"]
    for m in muster[:10]:
        lines.append(f"- '{m['name_pattern']}': {m['temp_bereich']}, "
                     f"Raumtemp {m.get('raum_temp_soll_c','?')}C, "
                     f"Komponenten: {', '.join(m.get('komponenten_erkannt',[]))}")
    return "\n".join(lines)

# --- SYSTEM PROMPT ---

def _sys():
    lk = lernmuster_als_kontext()
    return f"""Du bist ein erfahrener Kaeltetechnik-Planer bei coolsulting e.U. in Oesterreich.
Du analysierst Kaelteanlagenplaene, Ausschreibungen, Excel-Listen und Fotos.

{lk}

Erkenne alle Kuehlstellen und extrahiere:
- Name/Bezeichnung (exakt wie im Dokument – Pos.-Nr., Raumname, Code)
- Raumtemperatur Soll (C), Verdampfungstemperatur, Temperaturbereich (HNK/NK/NK+/TK/TK+)
- Kaelteleistung (kW), Lieferumfang (direkt/extern), Kreis-Nummer
- Position / Standort falls erkennbar
- Komponenten aus: verdampfer_luefter, abtauheizung, temperaturfuehler_innen,
  temperaturfuehler_abtau, eev, gehaeuse_heizung, tuer_heizung, tuer_kontakt,
  personennotruf, gaswarnanlage, monitoring_kiconex, schaltkasten, magnetventil, druckfuehler

KREISZUORDNUNG – SEHR WICHTIG:
Kuehlkreise (Refrigeration Circuits) muessen korrekt erkannt werden:
- Suche nach expliziten Kreisbezeichnungen: "Kreis 1", "Kreis NK", "Circuit 1", "Kreis HT/NT",
  "Gruppe A/B/C", "Sauggruppe", "Saug 1/2", "NK-Kreis", "TK-Kreis", "LT/MT/HT" etc.
- In Excel: oft eigene Spalte "Kreis", "Gruppe", "Circuit", "Saugseite"
- Im Plan: Rohrleitungen oder Farbcodierung zeigen Kreiszugehoerigkeit
- Kreise gruppieren sich IMMER nach aehnlicher Verdampfungstemperatur:
  HNK ca. +5C / NK ca. -8C / NK+ ca. -10C / TK ca. -31C bis -33C
- Wenn keine explizite Kreiszuordnung im Dokument: automatisch nach
  Verdampfungstemperatur gruppieren und nummerieren (1=waermster, aufsteigend)
- Gib IMMER eine Kreis-Nummer (1,2,3...) zurueck – nie null oder leer
- In "kreise" alle erkannten Kreise mit Bezeichnung aus dem Dokument zusammenfassen

NUR JSON ausgeben:
{{
  "projekt_info": {{"name": "", "standort": "", "bemerkung": ""}},
  "kreise": [
    {{"kreis_nr": 1, "bezeichnung": "NK-Kreis", "verdampfung_c": -8, "temp_bereich": "NK",
      "beschreibung": "Bezeichnung aus Dokument falls vorhanden"}}
  ],
  "kuehlstellen": [{{
    "nummer": 1, "name": "Pos. 1.1 Kuehlraum Fleisch", "beschreibung": "EG links",
    "temp_bereich": "NK", "raum_temp_soll_c": 0, "verdampfung_c": -8,
    "kaelteleistung_kw": 2.5, "kreis": 1, "lieferumfang": "direkt",
    "pos_nr": "1.1",
    "komponenten_erkannt": ["verdampfer_luefter", "abtauheizung"],
    "notizen": "laut Ausschreibung Pos. 116"
  }}],
  "hinweise": []
}}

WICHTIG - TEMPERATURREGELN:
- Immer den KAELTESTEN Wert verwenden, nicht den Mittelwert!
  Beispiel: 0-4 Grad -> verwende 0 Grad als raum_temp_soll_c
- Verdampfungstemperatur entsprechend anpassen (ca. 8-10K unter Raumtemp bei NK)
- HNK: Raumtemp bis +12C, Verdampfung ca. +2 bis +5C
- NK:  Raumtemp bis 0C,   Verdampfung ca. -8 bis -10C
- NK+: Raumtemp bis -5C,  Verdampfung ca. -10 bis -12C
- TK:  Raumtemp bis -22C, Verdampfung ca. -31 bis -33C
- TK+: Raumtemp unter -25C, Verdampfung ca. -33C und tiefer
- Luefteranzahl: Standard 1 – nur erhoehen wenn explizit mehrere erkennbar
- pos_nr: Positionsnummer aus Ausschreibung/Excel falls vorhanden (z.B. "116", "1.1", "A3")
direkt=unser Gewerk, extern=Fremdgewerk. Unsicherheiten in hinweise eintragen."""

# --- PDF ---

def analysiere_pdf(pdf_bytes):
    if not api_verfuegbar():
        return {"fehler": "Claude API nicht verfuegbar"}
    try:
        b64 = base64.standard_b64encode(pdf_bytes).decode("utf-8")
        r = _client().messages.create(
            model="claude-opus-4-5", max_tokens=4000, system=_sys(),
            messages=[{"role":"user","content":[
                {"type":"document","source":{"type":"base64","media_type":"application/pdf","data":b64}},
                {"type":"text","text":"Analysiere diesen Kaelteplan und extrahiere alle Kuehlstellen als JSON."}
            ]}])
        res = _parse(r.content[0].text)
        lerne_aus_ergebnis(res)
        return res
    except Exception as e:
        return {"fehler": str(e)}

# --- DWG ---

def extrahiere_dwg_text(dwg_bytes):
    if not EZDXF_OK:
        return "ezdxf nicht installiert"
    try:
        # Sicherstellen dass es Bytes sind
        if isinstance(dwg_bytes, str):
            dwg_bytes = dwg_bytes.encode("utf-8", errors="replace")

        # ezdxf kann nur DXF lesen, nicht binäres DWG
        # Versuche DXF-Format (ASCII oder Binary)
        try:
            doc = ezdxf.read(io.BytesIO(dwg_bytes))
        except Exception:
            # Fallback: als Text-DXF versuchen
            try:
                text_content = dwg_bytes.decode("utf-8", errors="ignore")
                doc = ezdxf.read(io.StringIO(text_content))
            except Exception as e2:
                return f"Fehler: Datei konnte nicht gelesen werden. Bitte als DXF oder PDF exportieren. ({str(e2)[:100]})"

        texte = []
        for e in doc.modelspace():
            if e.dxftype() in ("TEXT","MTEXT"):
                try:
                    t = e.dxf.text if e.dxftype()=="TEXT" else e.text
                    if t and len(t.strip())>1:
                        texte.append(t.strip())
                except Exception:
                    pass
            try:
                l = e.dxf.layer
                lbl = f"[Layer: {l}]"
                if l and lbl not in texte:
                    texte.append(lbl)
            except Exception:
                pass
        return "\n".join(texte) if texte else "Keine Texte gefunden – Plan enthält möglicherweise nur Grafiken"
    except Exception as e:
        return f"Fehler: {str(e)[:200]}"


def analysiere_dwg(dwg_bytes, dateiname):
    if not api_verfuegbar():
        return {"fehler": "Claude API nicht verfuegbar"}

    # DWG (binär) wird von ezdxf nicht unterstützt – Hinweis geben
    if dateiname.lower().endswith(".dwg"):
        # Trotzdem versuchen – manchmal sind DWG-Dateien intern DXF
        txt = extrahiere_dwg_text(dwg_bytes)
        if "Fehler:" in txt or "konnte nicht gelesen" in txt:
            return {
                "fehler": "DWG-Binärformat nicht lesbar. Bitte in AutoCAD/LibreCAD/cloudconvert.com "
                          "als DXF oder PDF exportieren und dann hochladen.",
                "tipp": "PDF-Export empfohlen: cloudconvert.com → DWG zu PDF → hier hochladen"
            }
    else:
        txt = extrahiere_dwg_text(dwg_bytes)

    if not txt or "nicht installiert" in txt:
        return {"fehler": txt}

    try:
        r = _client().messages.create(
            model="claude-opus-4-5", max_tokens=4000, system=_sys(),
            messages=[{"role":"user","content":
                f"DXF-Datei '{dateiname}' – extrahierter Text:\n\n{txt[:8000]}\n\nAlle Kuehlstellen als JSON."}])
        res = _parse(r.content[0].text)
        lerne_aus_ergebnis(res)
        return res
    except Exception as e:
        return {"fehler": str(e)}

# --- EXCEL ---

def extrahiere_excel_text(excel_bytes):
    if not EXCEL_OK:
        return "openpyxl nicht installiert"
    try:
        wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), data_only=True)
        out = []
        for sn in wb.sheetnames:
            ws = wb[sn]
            out.append(f"\n=== SHEET: {sn} ===")
            for row in ws.iter_rows(values_only=True):
                z = [str(c) if c is not None else "" for c in row]
                if any(x.strip() for x in z):
                    out.append(" | ".join(z))
        return "\n".join(out)
    except Exception as e:
        return f"Fehler: {str(e)}"

def analysiere_excel(excel_bytes, dateiname):
    if not api_verfuegbar():
        return {"fehler": "Claude API nicht verfuegbar"}
    txt = extrahiere_excel_text(excel_bytes)
    if "nicht installiert" in txt or "Fehler:" in txt:
        return {"fehler": txt}
    try:
        r = _client().messages.create(
            model="claude-opus-4-5", max_tokens=4000, system=_sys(),
            messages=[{"role":"user","content":
                f"Excel-Datei '{dateiname}':\n\n{txt[:8000]}\n\nAlle Kuehlstellen als JSON."}])
        res = _parse(r.content[0].text)
        lerne_aus_ergebnis(res)
        return res
    except Exception as e:
        return {"fehler": str(e)}

# --- BILD (Screenshot / Foto) ---

def analysiere_bild(bild_bytes, dateiname):
    if not api_verfuegbar():
        return {"fehler": "Claude API nicht verfuegbar"}
    fn = dateiname.lower()
    if fn.endswith(".png"):
        mt = "image/png"
    elif fn.endswith(".webp"):
        mt = "image/webp"
    else:
        mt = "image/jpeg"
    try:
        b64 = base64.standard_b64encode(bild_bytes).decode("utf-8")
        r = _client().messages.create(
            model="claude-opus-4-5", max_tokens=4000, system=_sys(),
            messages=[{"role":"user","content":[
                {"type":"image","source":{"type":"base64","media_type":mt,"data":b64}},
                {"type":"text","text":"Analysiere dieses Bild (Kaelteplan/Ausschreibung/Tabelle) und extrahiere alle Kuehlstellen als JSON."}
            ]}])
        res = _parse(r.content[0].text)
        lerne_aus_ergebnis(res)
        return res
    except Exception as e:
        return {"fehler": str(e)}

# --- FREITEXT ---

def analysiere_freitext(text):
    if not api_verfuegbar():
        return {"fehler": "Claude API nicht verfuegbar"}
    try:
        r = _client().messages.create(
            model="claude-opus-4-5", max_tokens=4000, system=_sys(),
            messages=[{"role":"user","content":
                f"Ausschreibung/Beschreibung:\n\n{text}\n\nAlle Kuehlstellen als JSON."}])
        res = _parse(r.content[0].text)
        lerne_aus_ergebnis(res)
        return res
    except Exception as e:
        return {"fehler": str(e)}

# --- JSON PARSER ---

def _parse(raw):
    """Robuster JSON-Parser."""
    if not raw:
        return {"fehler": "Leere Antwort"}
    # 1. Direkt
    try:
        return json.loads(raw)
    except Exception:
        pass
    # 2. Backtick-Bloecke
    for pat in [r"```json\s*([\s\S]*?)```", r"```\s*([\s\S]*?)```"]:
        m = re.search(pat, raw)
        if m:
            try: return json.loads(m.group(1).strip())
            except Exception: pass
    # 3. Groessten JSON-Block
    start = raw.find("{")
    end = raw.rfind("}")
    if start != -1 and end > start:
        candidate = raw[start:end+1]
        try: return json.loads(candidate)
        except Exception: pass
        # 4. Trailing commas fixen
        try:
            fixed = re.sub(r",\s*([}\]])", r"\1", candidate)
            return json.loads(fixed)
        except Exception: pass
    # 5. Nur kuehlstellen-Array
    km = re.search(r'"kuehlstellen"\s*:\s*(\[[\s\S]*?\])', raw)
    if km:
        try:
            return {"kuehlstellen": json.loads(km.group(1)),
                    "projekt_info": {}, "kreise": [],
                    "hinweise": ["Nur Kuehlstellen-Teil geparst"]}
        except Exception: pass
    return {"fehler": "JSON nicht parsbar", "raw": raw[:1000]}


def merge_kuehlstellen(bestehende: list, neue: list) -> tuple:
    """Fuehrt neue KI-Kuehlstellen mit bestehenden zusammen – erkennt Duplikate."""
    result = list(bestehende)
    neu_count = 0
    aktualisiert_count = 0
    for ks_neu in neue:
        pos_neu = str(ks_neu.get("pos_nr","")).strip().lower()
        name_neu = ks_neu.get("name","").strip().lower()
        gefunden_idx = None
        for i, ks_alt in enumerate(result):
            pos_alt = str(ks_alt.get("pos_nr","")).strip().lower()
            name_alt = ks_alt.get("name","").strip().lower()
            if pos_neu and pos_alt and pos_neu == pos_alt:
                gefunden_idx = i; break
            if len(name_neu) >= 5 and len(name_alt) >= 5:
                if name_neu[:15] in name_alt or name_alt[:15] in name_neu:
                    gefunden_idx = i; break
        if gefunden_idx is not None:
            alt = result[gefunden_idx]
            for key in ["temp_bereich","raum_temp_soll_c","verdampfung_custom_c",
                        "kaelteleistung_kw","kreis","beschreibung","notizen","pos_nr"]:
                if ks_neu.get(key) is not None:
                    alt[key] = ks_neu[key]
            for k, v in ks_neu.get("komponenten",{}).items():
                if k not in alt.get("komponenten",{}):
                    if "komponenten" not in alt: alt["komponenten"] = {}
                    alt["komponenten"][k] = v
            result[gefunden_idx] = alt
            aktualisiert_count += 1
        else:
            ks_neu["nummer"] = len(result) + 1
            result.append(ks_neu)
            neu_count += 1
    return result, neu_count, aktualisiert_count


# --- KONVERTIERUNG ---

def ki_ergebnis_zu_kuehlstellen(ki_result):
    from modules.kuehlstellen import neue_kuehlstelle, KOMPONENTEN
    ks_liste = []
    for i, ki_ks in enumerate(ki_result.get("kuehlstellen",[]), 1):
        ks = neue_kuehlstelle(i)
        ks["name"]              = ki_ks.get("name", f"Kuehlstelle {i}")
        ks["beschreibung"]      = ki_ks.get("beschreibung","")
        ks["temp_bereich"]      = ki_ks.get("temp_bereich","NK")
        ks["raum_temp_soll_c"]  = ki_ks.get("raum_temp_soll_c")
        ks["kaelteleistung_kw"] = ki_ks.get("kaelteleistung_kw")
        ks["kreis"]             = ki_ks.get("kreis", 1)
        ks["lieferumfang"]      = ki_ks.get("lieferumfang","direkt")
        ks["notizen"]           = ki_ks.get("notizen","")
        ks["pos_nr"]            = ki_ks.get("pos_nr","")
        if ki_ks.get("verdampfung_c"):
            ks["verdampfung_custom_c"] = ki_ks["verdampfung_c"]
        for komp_key in ki_ks.get("komponenten_erkannt",[]):
            if komp_key in KOMPONENTEN:
                ki_info = KOMPONENTEN[komp_key]
                ks["komponenten"][komp_key] = {
                    "aktiv": True,
                    "parameter": {pk:pv.get("vorschlag") for pk,pv in ki_info.get("parameter",{}).items()},
                    "ki_erkannt": True
                }
        ks_liste.append(ks)
    return ks_liste


def get_erkannte_kreise(ki_result: dict) -> list:
    """Gibt erkannte Kreise aus KI-Ergebnis zurück, mit Fallback aus Kühlstellen."""
    kreise = ki_result.get("kreise", [])
    if kreise:
        return kreise
    # Fallback: aus Kühlstellen ableiten
    from modules.kuehlstellen import TEMP_BEREICHE
    seen = {}
    for ks in ki_result.get("kuehlstellen", []):
        k_nr = ks.get("kreis", 1)
        if k_nr not in seen:
            tb = ks.get("temp_bereich", "NK")
            seen[k_nr] = {
                "kreis_nr": k_nr,
                "bezeichnung": f"Kreis {k_nr}",
                "verdampfung_c": ks.get("verdampfung_c") or TEMP_BEREICHE.get(tb,{}).get("verdampfung_c","?"),
                "temp_bereich": tb,
                "beschreibung": TEMP_BEREICHE.get(tb,{}).get("label","")
            }
    return sorted(seen.values(), key=lambda x: x["kreis_nr"])


# =============================================================================
# DATENBLATT-ANALYSE (Verdampfer, Außenunits)
# =============================================================================

def _sys_datenblatt():
    return """Du bist ein erfahrener Kältetechnik-Planer. Analysiere das hochgeladene Datenblatt
eines Kältegeräts (Verdampfer, Außenunit, Verflüssiger, Kompaktaggregat).

Extrahiere ALLE technischen Daten die du finden kannst. NUR JSON ausgeben:
{
  "typ": "Verdampfer",
  "hersteller": "Güntner",
  "modell": "GVHN 025.1D",
  "kaelteleistung_kw": 2.5,
  "kaeltemittel": "R513a",
  "temp_bereich": "NK",
  "raum_temp_c": 0,
  "verdampfung_c": -8,
  "anzahl_luefter": 1,
  "leistung_luefter_w": 140,
  "spannung_luefter": "230V 1-phasig",
  "strom_luefter_a": 0.7,
  "abtau_typ": "Elektro-Abtau",
  "abtau_leistung_w": 800,
  "abtau_spannung": "230V 1-phasig",
  "ablauf_heizung": false,
  "anschluss_spannung": "230V 1-phasig",
  "anschluss_strom_a": 2.5,
  "gewicht_kg": 18,
  "abmessungen_mm": "922x1003x266",
  "schutzart": "IP44",
  "freigabe_kontakt": true,
  "stoermeldung_kontakt": true,
  "betrieb_kontakt": false,
  "bus_schnittstelle": "RS485 Modbus",
  "steuerung_kabel_typ": "AU-7x1-STEUER",
  "notizen": "EC-Motoren, stufenlose Drehzahlregelung",
  "typ_erkannt_aus": "Produktdatenblatt Seite 1"
}

Typen: Verdampfer / Außenunit (Verflüssigungseinheit) / Verflüssiger (getrennt) / Kompaktaggregat / Verbundanlage
Temp-Bereich: HNK / NK / NK+ / TK / TK+
Abtau: Elektro-Abtau / Heißgas-Abtau / Umluft-Abtau / Keine
Wenn ein Wert nicht erkennbar: null setzen, nicht raten."""


def analysiere_datenblatt_pdf(pdf_bytes: bytes) -> dict:
    """Analysiert ein Gerätedatenblatt als PDF."""
    if not api_verfuegbar():
        return {"fehler": "Claude API nicht verfuegbar"}
    try:
        import base64
        b64 = base64.standard_b64encode(pdf_bytes).decode("utf-8")
        r = _client().messages.create(
            model="claude-opus-4-5", max_tokens=2000,
            system=_sys_datenblatt(),
            messages=[{"role":"user","content":[
                {"type":"document","source":{"type":"base64","media_type":"application/pdf","data":b64}},
                {"type":"text","text":"Analysiere dieses Gerätedatenblatt und extrahiere alle technischen Daten als JSON."}
            ]}])
        return _parse(r.content[0].text)
    except Exception as e:
        return {"fehler": str(e)}


def analysiere_datenblatt_bild(img_bytes: bytes, dateiname: str = "") -> dict:
    """Analysiert ein Gerätedatenblatt als Bild/Screenshot."""
    if not api_verfuegbar():
        return {"fehler": "Claude API nicht verfuegbar"}
    try:
        import base64
        ext = dateiname.lower().split(".")[-1] if "." in dateiname else "jpeg"
        mime = {"jpg":"image/jpeg","jpeg":"image/jpeg","png":"image/png","webp":"image/webp"}.get(ext,"image/jpeg")
        b64 = base64.standard_b64encode(img_bytes).decode("utf-8")
        r = _client().messages.create(
            model="claude-opus-4-5", max_tokens=2000,
            system=_sys_datenblatt(),
            messages=[{"role":"user","content":[
                {"type":"image","source":{"type":"base64","media_type":mime,"data":b64}},
                {"type":"text","text":"Analysiere dieses Gerätedatenblatt/Produktbild und extrahiere alle technischen Daten als JSON."}
            ]}])
        return _parse(r.content[0].text)
    except Exception as e:
        return {"fehler": str(e)}


def analysiere_datenblatt_excel(excel_bytes: bytes) -> dict:
    """Analysiert eine Excel-Stückliste/Geräteliste."""
    if not api_verfuegbar():
        return {"fehler": "Claude API nicht verfuegbar"}
    try:
        import io
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), read_only=True)
        text_parts = []
        for ws in wb.worksheets[:3]:
            rows = []
            for row in ws.iter_rows(max_row=100, values_only=True):
                vals = [str(c) if c is not None else "" for c in row]
                if any(v.strip() for v in vals):
                    rows.append(" | ".join(vals))
            if rows:
                text_parts.append("Sheet: " + ws.title + "\n" + "\n".join(rows[:80]))
        text = "\n\n".join(text_parts)[:6000]
        r = _client().messages.create(
            model="claude-opus-4-5", max_tokens=2000,
            system=_sys_datenblatt(),
            messages=[{"role":"user","content":
                "Geräteliste/Stückliste aus Excel:\n\n" + text + "\n\nExtrahiere Gerätedaten als JSON-Array mit einem Objekt pro Gerät."}])
        result = _parse(r.content[0].text)
        # Wenn Array zurückkommt, erstes Gerät nehmen oder als Liste behandeln
        if isinstance(result, list):
            return {"geraete": result}
        return result
    except Exception as e:
        return {"fehler": str(e)}
